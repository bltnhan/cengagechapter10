import streamlit as st
import pandas as pd
import numpy as np
import io
import json
import re as _re
import warnings
from collections import Counter
warnings.filterwarnings('ignore')

st.set_page_config(page_title="DataMine AI", page_icon="🧬", layout="wide",
                   initial_sidebar_state="expanded")

# ═══════════════════════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;700&display=swap');

html,body,[class*="css"]{font-family:'Inter',sans-serif;}
.stApp{background:#0a0e1a;color:#f1f5f9;}
.main .block-container{padding:1.2rem 2rem;max-width:1400px;}
::-webkit-scrollbar{width:5px;height:5px;}
::-webkit-scrollbar-track{background:#111827;}
::-webkit-scrollbar-thumb{background:#374151;border-radius:3px;}

[data-testid="stSidebar"]{background:#0d1120!important;border-right:1px solid #1f2937;}
h1,h2,h3,h4{color:#f9fafb!important;font-family:'Inter',sans-serif!important;}
[data-testid="stMetricValue"]{color:#f9fafb!important;font-weight:700!important;}
[data-testid="stMetricLabel"]{color:#9ca3af!important;}
.stDataFrame td,.stDataFrame th{color:#e5e7eb!important;background:#111827!important;font-size:.82rem!important;}
.stSelectbox div[data-baseweb="select"],.stMultiSelect div[data-baseweb="select"]{background:#111827!important;color:#f1f5f9!important;border-color:#374151!important;}
div[data-baseweb="option"]{background:#111827!important;color:#f1f5f9!important;}
div[data-baseweb="popover"]{background:#111827!important;border:1px solid #374151!important;}
.stTextInput input,.stTextArea textarea{color:#f1f5f9!important;background:#111827!important;border-color:#374151!important;}
button[data-baseweb="tab"]{color:#9ca3af!important;font-family:'Inter',sans-serif!important;}
button[data-baseweb="tab"][aria-selected="true"]{color:#60a5fa!important;}
[data-testid="stFileUploadDropzone"]{background:#111827!important;border:2px dashed #374151!important;border-radius:12px!important;}
[data-testid="stFileUploadDropzone"]:hover{border-color:#3b82f6!important;}
[data-baseweb="tab-list"]{background:#111827!important;border-radius:8px!important;padding:4px!important;border:1px solid #1f2937!important;}
[data-baseweb="tab"][aria-selected="true"]{background:#1d4ed8!important;}
details{background:#111827!important;border:1px solid #1f2937!important;border-radius:10px!important;}
summary{color:#9ca3af!important;}
.stInfo{background:rgba(59,130,246,.08)!important;border-left-color:#3b82f6!important;}
.stSuccess{background:rgba(16,185,129,.08)!important;border-left-color:#10b981!important;}
.stWarning{background:rgba(245,158,11,.08)!important;border-left-color:#f59e0b!important;}
.stError{background:rgba(239,68,68,.08)!important;border-left-color:#ef4444!important;}

.stButton>button{background:linear-gradient(135deg,#2563eb,#3b82f6)!important;color:white!important;
  border:none!important;border-radius:8px!important;font-family:'Inter',sans-serif!important;
  font-weight:600!important;padding:.45rem 1.3rem!important;transition:all .2s!important;}
.stButton>button:hover{opacity:.85!important;}
.stDownloadButton>button{background:linear-gradient(135deg,#065f46,#10b981)!important;}

.app-header{display:flex;align-items:center;gap:14px;padding:0 0 1.2rem 0;margin-bottom:.8rem;border-bottom:1px solid #1f2937;}
.app-logo{font-size:1.6rem;width:44px;height:44px;background:linear-gradient(135deg,#1d4ed8,#7c3aed);
  border-radius:11px;display:flex;align-items:center;justify-content:center;}
.app-title{font-size:1.45rem;font-weight:700;color:#f9fafb;letter-spacing:-.02em;}
.app-sub{font-size:.78rem;color:#6b7280;margin-top:1px;}
.version-badge{margin-left:auto;background:linear-gradient(135deg,#1d4ed8,#7c3aed);color:white;
  font-size:.68rem;font-weight:700;padding:3px 11px;border-radius:20px;letter-spacing:.05em;}

.stepper{display:flex;align-items:center;background:#111827;border:1px solid #1f2937;
  border-radius:14px;padding:14px 24px;margin-bottom:1.5rem;}
.step-item{display:flex;flex-direction:column;align-items:center;gap:5px;flex:0 0 auto;}
.step-circle{width:34px;height:34px;border-radius:50%;display:flex;align-items:center;
  justify-content:center;font-size:.78rem;font-weight:700;transition:all .3s;}
.sc-p{background:#1f2937;color:#4b5563;border:2px solid #374151;}
.sc-a{background:linear-gradient(135deg,#1d4ed8,#3b82f6);color:#fff;border:2px solid #3b82f6;box-shadow:0 0 14px rgba(59,130,246,.4);}
.sc-d{background:linear-gradient(135deg,#065f46,#10b981);color:#fff;border:2px solid #10b981;}
.step-lbl{font-size:.68rem;font-weight:600;letter-spacing:.03em;white-space:nowrap;}
.sl-p{color:#4b5563;}.sl-a{color:#60a5fa;}.sl-d{color:#34d399;}
.step-conn{flex:1;height:2px;margin:0 6px;margin-bottom:18px;border-radius:1px;}
.sc-conn-d{background:linear-gradient(90deg,#10b981,#059669);}
.sc-conn-p{background:#1f2937;}

.stat-row{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:1rem;}
.stat-card{background:#111827;border:1px solid #1f2937;border-radius:11px;padding:14px 18px;}
.stat-val{font-size:1.5rem;font-weight:700;font-family:'JetBrains Mono',monospace;}
.stat-lbl{font-size:.72rem;color:#6b7280;margin-top:3px;letter-spacing:.04em;text-transform:uppercase;}
.sv-blue{color:#60a5fa;}.sv-green{color:#34d399;}.sv-red{color:#f87171;}.sv-amber{color:#fbbf24;}.sv-gray{color:#9ca3af;}

.sec-hdr{display:flex;align-items:center;gap:10px;margin:1.4rem 0 .9rem;padding-bottom:9px;border-bottom:1px solid #1f2937;}
.sec-icon{width:30px;height:30px;border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:.9rem;}
.si-blue{background:rgba(59,130,246,.15);}.si-green{background:rgba(16,185,129,.15);}
.si-purple{background:rgba(124,58,237,.15);}.si-amber{background:rgba(245,158,11,.15);}
.sec-title{font-size:.92rem;font-weight:700;color:#f1f5f9;}
.sec-sub{font-size:.72rem;color:#6b7280;margin-left:auto;}

.card{background:#111827;border:1px solid #1f2937;border-radius:11px;padding:16px 18px;margin-bottom:10px;}

.prob-row{display:flex;align-items:flex-start;gap:8px;padding:8px 11px;border-radius:8px;margin-bottom:5px;background:#0f1629;border:1px solid #1f2937;}
.prob-chip{padding:2px 8px;border-radius:5px;font-size:.67rem;font-weight:700;letter-spacing:.06em;white-space:nowrap;font-family:'JetBrains Mono',monospace;}
.pc-err{background:rgba(239,68,68,.2);color:#f87171;}
.pc-warn{background:rgba(245,158,11,.2);color:#fbbf24;}
.pc-info{background:rgba(59,130,246,.2);color:#60a5fa;}
.pc-ok{background:rgba(16,185,129,.2);color:#34d399;}
.prob-txt{font-size:.81rem;color:#d1d5db;flex:1;}
.prob-fix{font-size:.72rem;color:#6b7280;font-style:italic;}

.method-card{background:#111827;border:1px solid #1f2937;border-radius:10px;padding:11px 13px;margin-bottom:4px;}
.method-card.sel{border-color:#10b981;background:#052e1c;box-shadow:0 0 0 1px #10b981;}
.method-name{font-size:.83rem;font-weight:600;color:#f1f5f9;}
.method-vn{font-size:.7rem;color:#6b7280;margin-top:1px;}
.method-desc{font-size:.71rem;color:#9ca3af;margin-top:5px;line-height:1.4;}
.mbadge{display:inline-block;padding:1px 7px;border-radius:4px;font-size:.65rem;font-weight:700;margin-bottom:3px;}
.mb-clf{background:rgba(59,130,246,.2);color:#60a5fa;}
.mb-pred{background:rgba(124,58,237,.2);color:#a78bfa;}
.mb-clus{background:rgba(16,185,129,.2);color:#34d399;}
.mb-assoc{background:rgba(239,68,68,.2);color:#f87171;}
.mb-bal{background:rgba(245,158,11,.2);color:#fbbf24;}

.ai-box{background:#0d1829;border:1px solid #1e3a5f;border-radius:12px;padding:16px 18px;margin:10px 0;}
.ai-hdr{font-size:.78rem;font-weight:700;color:#60a5fa;margin-bottom:10px;display:flex;align-items:center;gap:7px;}
.ai-text{color:#d1d5db;font-size:.85rem;line-height:1.8;white-space:pre-wrap;word-break:break-word;}
.ai-dot{width:7px;height:7px;background:#3b82f6;border-radius:50%;animation:pulse 2s infinite;}
@keyframes pulse{0%,100%{opacity:1;}50%{opacity:.3;}}

.log-item{background:rgba(16,185,129,.05);border-left:3px solid #10b981;padding:5px 11px;border-radius:0 6px 6px 0;margin-bottom:4px;font-size:.79rem;color:#d1d5db;}
.log-item b{color:#34d399;}
.diff-before{background:rgba(239,68,68,.05);border:1px solid rgba(239,68,68,.25);border-radius:9px;padding:10px 14px;}
.diff-after{background:rgba(16,185,129,.05);border:1px solid rgba(16,185,129,.25);border-radius:9px;padding:10px 14px;}
.diff-label{font-size:.75rem;font-weight:700;margin-bottom:6px;}
.dl-red{color:#f87171;}.dl-green{color:#34d399;}

.sb-step{display:flex;align-items:center;gap:9px;padding:7px 10px;border-radius:8px;margin-bottom:2px;}
.sb-step-a{background:rgba(59,130,246,.12);}
.sb-num{width:20px;height:20px;border-radius:50%;font-size:.68rem;font-weight:700;display:flex;align-items:center;justify-content:center;}
.sn-done{background:rgba(16,185,129,.2);color:#34d399;}
.sn-active{background:rgba(59,130,246,.2);color:#60a5fa;}
.sn-pend{background:#1f2937;color:#4b5563;}
.sb-txt{font-size:.8rem;font-weight:500;}
.sbt-a{color:#93c5fd;}.sbt-d{color:#6ee7b7;}.sbt-p{color:#4b5563;}

.chart-title{font-size:.82rem;font-weight:600;color:#e5e7eb;margin-bottom:4px;}
.chart-cap{font-size:.71rem;color:#6b7280;margin-top:3px;font-style:italic;}

.welcome{background:linear-gradient(135deg,#0f1f3d,#150f2d);border:1px solid #1e3a5f;border-radius:14px;padding:28px 32px;margin-bottom:1.5rem;}
.welcome h2{font-size:1.6rem!important;background:linear-gradient(90deg,#60a5fa,#a78bfa,#f472b6);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:8px!important;}
.welcome p{color:#9ca3af;font-size:.88rem;line-height:1.7;}
.step-list{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin-top:16px;}
.step-list-item{background:rgba(255,255,255,.03);border:1px solid #1f2937;border-radius:9px;padding:12px 14px;display:flex;align-items:flex-start;gap:10px;}
.sli-num{font-size:.75rem;font-weight:700;color:#60a5fa;font-family:'JetBrains Mono',monospace;background:rgba(59,130,246,.1);border-radius:5px;padding:2px 7px;white-space:nowrap;}
.sli-text{font-size:.8rem;color:#9ca3af;line-height:1.5;}
.sli-text b{color:#e5e7eb;display:block;margin-bottom:2px;}

/* ── Smart Advisor ── */
.quiz-card{background:#0f1629;border:1px solid #1e3a5f;border-radius:11px;padding:14px 18px;margin-bottom:10px;}
.quiz-q{font-size:.87rem;font-weight:600;color:#f1f5f9;margin-bottom:8px;}
.quiz-hint{font-size:.72rem;color:#6b7280;margin-top:-5px;margin-bottom:9px;}
.decision-path{display:flex;align-items:center;gap:5px;flex-wrap:wrap;margin:8px 0;padding:9px 12px;background:#0a0e1a;border-radius:8px;border:1px solid #1f2937;}
.dp-node{padding:3px 9px;border-radius:5px;font-size:.71rem;font-weight:600;}
.dp-blue{background:rgba(59,130,246,.15);color:#60a5fa;}.dp-green{background:rgba(16,185,129,.15);color:#34d399;}
.dp-purple{background:rgba(124,58,237,.15);color:#a78bfa;}.dp-amber{background:rgba(245,158,11,.15);color:#fbbf24;}
.dp-arrow{color:#374151;font-size:.8rem;}
.rec-item{display:flex;align-items:flex-start;gap:9px;padding:8px 10px;border-radius:8px;margin-bottom:5px;}
.ri-prereq{background:rgba(245,158,11,.07);border:1px solid rgba(245,158,11,.22);}
.ri-primary{background:rgba(59,130,246,.07);border:1px solid rgba(59,130,246,.22);}
.ri-alt{background:rgba(124,58,237,.06);border:1px solid rgba(124,58,237,.18);}
.ri-opt{background:rgba(16,185,129,.05);border:1px solid rgba(16,185,129,.15);}
.rec-badge{padding:2px 8px;border-radius:5px;font-size:.64rem;font-weight:700;white-space:nowrap;flex-shrink:0;}
.rb-pre{background:rgba(245,158,11,.2);color:#fbbf24;}.rb-pri{background:rgba(59,130,246,.2);color:#60a5fa;}
.rb-alt{background:rgba(124,58,237,.2);color:#a78bfa;}.rb-opt{background:rgba(16,185,129,.2);color:#34d399;}
.rec-name{font-size:.83rem;font-weight:600;color:#f1f5f9;}
.rec-why{font-size:.72rem;color:#9ca3af;margin-top:2px;line-height:1.4;}

/* ── Category cards ── */
.cat-card{border-radius:11px;padding:14px 16px;margin-bottom:8px;}

/* ── Fitness Report ── */
.fitness-card{background:#111827;border:1px solid #1f2937;border-radius:12px;padding:16px 18px;}
.fitness-grade{font-size:2.2rem;font-weight:900;font-family:'JetBrains Mono',monospace;line-height:1;}
.fg-A{color:#34d399;}.fg-B{color:#60a5fa;}.fg-C{color:#fbbf24;}.fg-D{color:#f87171;}
.fitness-method{font-size:.85rem;font-weight:700;color:#f1f5f9;margin:4px 0 2px;}
.fitness-score{font-size:.77rem;color:#6b7280;margin-bottom:8px;font-family:'JetBrains Mono',monospace;}
.fitness-verdict{font-size:.77rem;padding:5px 10px;border-radius:6px;margin-bottom:8px;font-weight:600;}
.fv-A{background:rgba(16,185,129,.1);color:#34d399;border:1px solid rgba(16,185,129,.25);}
.fv-B{background:rgba(59,130,246,.1);color:#60a5fa;border:1px solid rgba(59,130,246,.25);}
.fv-C{background:rgba(245,158,11,.1);color:#fbbf24;border:1px solid rgba(245,158,11,.25);}
.fv-D{background:rgba(239,68,68,.1);color:#f87171;border:1px solid rgba(239,68,68,.25);}
.fitness-bar-wrap{background:#1f2937;border-radius:4px;height:5px;margin:5px 0 8px;}
.fitness-bar{border-radius:4px;height:5px;transition:width .5s ease;}
.fb-A{background:linear-gradient(90deg,#10b981,#34d399);}
.fb-B{background:linear-gradient(90deg,#2563eb,#60a5fa);}
.fb-C{background:linear-gradient(90deg,#d97706,#fbbf24);}
.fb-D{background:linear-gradient(90deg,#dc2626,#f87171);}
.fitness-advice{font-size:.75rem;color:#9ca3af;line-height:1.55;}

/* Thinking animation */
@keyframes thinking-blink {
  0%,100%{opacity:.2} 33%{opacity:1} 66%{opacity:.5}
}
.thinking-dots {
  display:inline-block;
  letter-spacing:2px;
  animation: thinking-blink 1.4s infinite;
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# ML IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════
import google.generativeai as genai
import requests as _requests
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier, export_text
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier
from sklearn.neural_network import MLPClassifier, MLPRegressor
from sklearn.cluster import KMeans, AgglomerativeClustering
from sklearn.decomposition import PCA
from sklearn.metrics import (accuracy_score, classification_report,
                             mean_squared_error, r2_score, silhouette_score,
                             confusion_matrix, precision_score, recall_score,
                             f1_score, roc_auc_score, roc_curve)
from mlxtend.frequent_patterns import apriori, association_rules
from mlxtend.preprocessing import TransactionEncoder
# imbalanced-learn không tương thích Python 3.14 — dùng implementation tự viết
try:
    from imblearn.over_sampling import RandomOverSampler, SMOTE
    _IMBLEARN_OK = True
except Exception:
    _IMBLEARN_OK = False

    class RandomOverSampler:
        """Fallback: random oversampling bằng numpy."""
        def __init__(self, random_state=42):
            self.random_state = random_state
        def fit_resample(self, X, y):
            rng = np.random.RandomState(self.random_state)
            classes, counts = np.unique(y, return_counts=True)
            max_count = counts.max()
            X_res, y_res = list(X), list(y)
            for cls, cnt in zip(classes, counts):
                if cnt < max_count:
                    idx = np.where(y == cls)[0]
                    extra = rng.choice(idx, max_count - cnt, replace=True)
                    X_res.extend(X[extra])
                    y_res.extend(y[extra])
            return np.array(X_res), np.array(y_res)

    class SMOTE:
        """Fallback: dùng RandomOverSampler khi imblearn không có."""
        def __init__(self, random_state=42):
            self._ros = RandomOverSampler(random_state=random_state)
        def fit_resample(self, X, y):
            return self._ros.fit_resample(X, y)
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.cluster.hierarchy import dendrogram, linkage

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
_DEFAULT_GEMINI_KEY = ""
_GEMINI_CANDIDATES  = ["gemini-2.5-flash","gemini-2.0-flash","gemini-1.5-flash","gemini-1.5-pro"]
_GEMINI_BASE_URL    = "https://generativelanguage.googleapis.com/v1beta/models"
_OPENROUTER_URL     = "https://openrouter.ai/api/v1/chat/completions"
_OPENROUTER_MODELS  = ["openrouter/auto","google/gemma-3-27b-it:free",
    "meta-llama/llama-3.3-70b-instruct:free","deepseek/deepseek-r1:free","qwen/qwq-32b:free"]

METHODS = {
    "Logistic Regression":{"group":"classification","badge":"mb-clf","vn":"Phân loại cơ bản",
        "desc":"Probabilistic S-curve classifier. Interpretable coefficients. Great baseline."},
    "Linear Discriminant Analysis (LDA)":{"group":"classification","badge":"mb-clf","vn":"Phân tích biệt thức",
        "desc":"Maximises class separation. Works best with equal class covariance."},
    "K-Nearest Neighbors (KNN)":{"group":"classification","badge":"mb-clf","vn":"Phân loại láng giềng",
        "desc":"Labels from K closest samples. Non-parametric, interpretable."},
    "Classification Trees":{"group":"classification","badge":"mb-clf","vn":"Cây quyết định",
        "desc":"Builds IF-THEN rules to split data. Human-readable output."},
    "Naive Bayes":{"group":"classification","badge":"mb-clf","vn":"Dự báo xác suất",
        "desc":"Bayes theorem with independence assumption. Very fast on sparse data."},
    "Support Vector Machine (SVM)":{"group":"classification","badge":"mb-clf","vn":"Phân loại biên giới",
        "desc":"Maximum-margin hyperplane. Powerful in high-dimensional spaces."},
    "Random Forest":{"group":"classification","badge":"mb-clf","vn":"Rừng ngẫu nhiên",
        "desc":"Ensemble of 100 trees. Robust, accurate, feature importance ranking."},
    "Neural Networks (MLP)":{"group":"classification","badge":"mb-clf","vn":"Mạng nơ-ron",
        "desc":"Multi-layer perceptron for complex non-linear patterns."},
    "Linear Regression":{"group":"prediction","badge":"mb-pred","vn":"Hồi quy tuyến tính",
        "desc":"Linear relationship between features and continuous target."},
    "Neural Networks Regression (MLP)":{"group":"prediction","badge":"mb-pred","vn":"Mạng nơ-ron hồi quy",
        "desc":"MLP for continuous value prediction via non-linear transformations."},
    "Association Rules (Apriori)":{"group":"association","badge":"mb-assoc","vn":"Luật kết hợp",
        "desc":"IF-THEN patterns in transactional data. Support, confidence, lift."},
    "K-Means Clustering":{"group":"association","badge":"mb-clus","vn":"Phân cụm K-Means",
        "desc":"Groups data into K clusters by minimising intra-cluster variance."},
    "Hierarchical Clustering":{"group":"association","badge":"mb-clus","vn":"Phân cụm phân cấp",
        "desc":"Dendrogram of nested clusters. No K needed upfront."},
    "Random Oversampling":{"group":"balancing","badge":"mb-bal","vn":"Cân bằng ngẫu nhiên",
        "desc":"Replicates minority samples to fix class imbalance."},
    "SMOTE":{"group":"balancing","badge":"mb-bal","vn":"Cân bằng tổng hợp",
        "desc":"Synthetic minority samples by interpolation between existing ones."},
}
GROUP_META = {
    "classification":{"label":"Classification","color":"#60a5fa","icon":"🔵"},
    "prediction":    {"label":"Prediction / Regression","color":"#a78bfa","icon":"🟣"},
    "association":   {"label":"Clustering & Association","color":"#f472b6","icon":"🔴"},
    # "balancing" group removed — handled by Class balancing dropdown instead
}

# ═══════════════════════════════════════════════════════════════════════════════
# BACKEND HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _df_to_json(df):
    buf=io.StringIO(); df.to_json(buf,orient="split"); return buf.getvalue()

def _json_to_df(s):
    return pd.read_json(io.StringIO(s),orient="split")

def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Convert tất cả column về dtype PyArrow-compatible. Không điều kiện."""
    df = df.copy()
    df.columns = [str(c) for c in df.columns]
    for col in df.columns:
        try:
            dtype = df[col].dtype
            # Extension types → unwrap về object trước
            if pd.api.types.is_extension_array_dtype(dtype):
                df[col] = df[col].astype(object)
                dtype = df[col].dtype
            # Object → str vô điều kiện (handles str+int+float mixed)
            if dtype == object:
                df[col] = df[col].fillna("").astype(str)
                df[col] = df[col].replace("nan", "").replace("<NA>", "")
            # Non-standard float → float64
            elif pd.api.types.is_float_dtype(dtype) and dtype != np.float64:
                df[col] = df[col].astype(np.float64)
            # Non-standard int → int64
            elif pd.api.types.is_integer_dtype(dtype) and dtype != np.int64:
                df[col] = df[col].astype(np.int64)
        except Exception:
            try:
                df[col] = df[col].fillna("").astype(str)
            except Exception:
                df[col] = ""
    return df

def _pick_best_sheet(sheet_names):
    """Chọn sheet tốt nhất: ưu tiên 'Data', bỏ qua 'Description'."""
    skip = {"description","instructions","notes","readme","info","about","legend"}
    preferred = {"data","sheet1","serving data","dataset","raw data","input","main","records"}
    for s in sheet_names:
        if s.lower() in preferred: return s
    for s in sheet_names:
        if s.lower() not in skip: return s
    return sheet_names[0]

def load_sheet(raw_bytes, sheet_name, engine_xl, header_row=0):
    """Đọc 1 sheet cụ thể từ bytes, trả về DataFrame đã sanitize."""
    df = pd.read_excel(io.BytesIO(raw_bytes), sheet_name=sheet_name,
                       header=header_row, engine=engine_xl)
    return _sanitize_df(df)

def load_file(uploaded, header_row: int = 0, sheet_name: str = None):
    """
    Đọc file upload — chỉ load 1 sheet (sheet được chọn hoặc sheet tốt nhất).
    Trả về: (key, df, sheet_names_list)  — sheet_names_list chỉ có với Excel
    """
    name = uploaded.name
    result = {}          # {display_key: df}
    all_sheet_names = [] # danh sách sheet có trong file (Excel only)
    try:
        if name.endswith((".csv", ".txt")):
            sep = "," if name.endswith(".csv") else None
            uploaded.seek(0)
            df = pd.read_csv(uploaded, sep=sep, header=header_row,
                             engine="python" if name.endswith(".txt") else "c")
            result[name] = _sanitize_df(df)

        elif name.endswith((".xlsx", ".xls", ".xlsm")):
            uploaded.seek(0)
            raw_bytes = uploaded.read()
            engine_xl = "xlrd" if name.endswith(".xls") else "openpyxl"
            xf = pd.ExcelFile(io.BytesIO(raw_bytes), engine=engine_xl)
            all_sheet_names = xf.sheet_names

            # Chọn sheet: dùng sheet_name nếu có, nếu không thì tự chọn tốt nhất
            chosen = sheet_name if (sheet_name and sheet_name in all_sheet_names) \
                     else _pick_best_sheet(all_sheet_names)

            df = load_sheet(raw_bytes, chosen, engine_xl, header_row)
            key = f"{name}::{chosen}"
            result[key] = df

            # Lưu raw_bytes + engine vào session_state để sau có thể đổi sheet
            st.session_state.setdefault("_file_cache", {})[name] = {
                "raw_bytes": raw_bytes, "engine": engine_xl,
                "sheet_names": all_sheet_names, "current_sheet": chosen
            }

        elif name.endswith(".json"):
            uploaded.seek(0)
            result[name] = _sanitize_df(pd.read_json(uploaded))

        else:
            uploaded.seek(0)
            result[name] = _sanitize_df(pd.read_csv(uploaded, sep=None, engine="python"))

    except Exception as e:
        st.error(f"❌ Lỗi đọc {name}: {e}")
    return result

def df_summary(df,name=""):
    lines=[f"File: {name}" if name else ""]
    lines.append(f"Shape: {df.shape[0]} rows x {df.shape[1]} columns")
    lines.append(f"Columns: {', '.join(df.columns.tolist())}")
    lines.append(f"Missing: {df.isnull().sum().sum()} total ({df.isnull().mean().mean():.1%})")
    lines.append(f"Dtypes: {dict(df.dtypes.value_counts())}")
    lines.append("Sample stats:"); lines.append(df.describe(include="all").to_string())
    return "\n".join(lines)

def encode_df(df):
    df=df.copy(); le=LabelEncoder()
    for col in df.select_dtypes(include=["object","string"]).columns:
        df[col]=le.fit_transform(df[col].astype(str))
    return df

def _fig_to_png(fig):
    """Chuyển matplotlib figure → PNG bytes để lưu vào session_state."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight",
                facecolor=fig.get_facecolor(), dpi=120)
    buf.seek(0)
    return buf.read()

def fig_to_st(fig, save_key=None):
    """Render figure lên Streamlit và optionally lưu bytes vào session_state."""
    if save_key:
        st.session_state.setdefault("_run_figures", {}).setdefault(save_key, []).append(
            _fig_to_png(fig)
        )
    st.pyplot(fig)
    plt.close(fig)

def _clean_ai(text):
    t=_re.sub(r"\*{1,3}(.*?)\*{1,3}",r"\1",text)
    t=_re.sub(r"^#+\s*","",t,flags=_re.MULTILINE)
    t=_re.sub(r"^[-•]\s+","  ",t,flags=_re.MULTILINE)
    t=_re.sub(r"`{1,3}","",t); t=_re.sub(r"\n{3,}","\n\n",t)
    return t.strip()

def _safe_html(s): return s.replace("<","&lt;").replace(">","&gt;")

def render_ai_box(text,label="🤖 AI Analysis"):
    clean=_clean_ai(text)
    st.markdown(
        f'<div class="ai-box"><div class="ai-hdr"><div class="ai-dot"></div>{label}</div>'
        f'<div class="ai-text">{_safe_html(clean)}</div></div>',
        unsafe_allow_html=True)

def _get_keys():
    # Pull from sidebar input first, fall back to Streamlit Cloud secrets
    g=st.session_state.get("gemini_key","").strip()
    o=st.session_state.get("openrouter_key","").strip()
    if not g:
        try: g=st.secrets.get("GEMINI_KEY","")
        except Exception: g=""
    if not o:
        try: o=st.secrets.get("OPENROUTER_KEY","")
        except Exception: o=""
    return g or _DEFAULT_GEMINI_KEY,o

def _call_openrouter(prompt,or_key):
    if not or_key: return ""
    headers={"Authorization":f"Bearer {or_key}","Content-Type":"application/json",
             "HTTP-Referer":"https://datamine-ai.streamlit.app","X-Title":"DataMine AI"}
    for model in _OPENROUTER_MODELS:
        try:
            resp=_requests.post(_OPENROUTER_URL,headers=headers,
                json={"model":model,"messages":[{"role":"user","content":prompt}],
                      "max_tokens":2000,"temperature":0.7},timeout=120)
            if resp.status_code==401: return "__OR_FAIL__: Invalid key."
            data=resp.json()
            if "choices" in data and data["choices"]:
                return data["choices"][0]["message"]["content"]
        except Exception: continue
    return ""

def _call_gemini_rest(prompt_text, api_key, model="gemini-2.5-flash"):
    """Gọi Gemini REST API trực tiếp — không phụ thuộc version thư viện."""
    url = f"{_GEMINI_BASE_URL}/{model}:generateContent"
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt_text}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 8192}
    }
    resp = _requests.post(url,
        headers={"Content-Type": "application/json", "x-goog-api-key": api_key},
        json=payload, timeout=60)
    if resp.status_code == 401:
        raise ValueError(f"API Key không hợp lệ (401). Kiểm tra lại key trong sidebar.")
    if resp.status_code == 429:
        raise ValueError(f"Rate limit (429). Chờ 60 giây rồi thử lại.")
    if resp.status_code == 404:
        raise ValueError(f"Model '{model}' không tồn tại (404).")
    resp.raise_for_status()
    data = resp.json()
    return data["candidates"][0]["content"]["parts"][0]["text"]

def ask_ai(prompt):
    gemini_key, or_key = _get_keys()
    if not gemini_key and not or_key:
        return "⚠️ Chưa nhập API Key. Vui lòng nhập Gemini key trong sidebar."
    last_err = None
    if gemini_key:
        for model_name in _GEMINI_CANDIDATES:
            try:
                return _call_gemini_rest(prompt, gemini_key, model_name)
            except ValueError as e:
                # Lỗi xác định (key sai, rate limit) — dừng luôn
                return f"❌ {e}"
            except Exception as e:
                last_err = e
                continue  # thử model tiếp
    if or_key:
        result = _call_openrouter(prompt, or_key)
        if result and not result.startswith("__OR_FAIL__"):
            return result
    err_str = str(last_err) if last_err else "Không tìm thấy lỗi cụ thể"
    return f"❌ Tất cả model đều thất bại.\nLỗi cuối: {err_str}"

def ask_ai_stream(messages_for_api):
    """Gọi Gemini SSE streaming — yield từng chunk text."""
    gemini_key, or_key = _get_keys()
    if not gemini_key and not or_key:
        raise ValueError("Chưa nhập API Key. Nhập Gemini key trong sidebar.")

    last_error = None

    if gemini_key:
        for model_name in _GEMINI_CANDIDATES:
            try:
                url = f"{_GEMINI_BASE_URL}/{model_name}:streamGenerateContent?alt=sse"
                payload = {
                    "contents": messages_for_api,
                    "generationConfig": {"temperature": 0.3, "maxOutputTokens": 8192}
                }
                resp = _requests.post(url,
                    headers={"Content-Type":"application/json","x-goog-api-key": gemini_key},
                    json=payload, stream=True, timeout=120)
                if resp.status_code == 401:
                    raise ValueError("API Key không hợp lệ (401). Kiểm tra lại key trong sidebar.")
                if resp.status_code == 429:
                    raise ValueError("Rate limit (429). Chờ 60 giây rồi thử lại.")
                if resp.status_code == 404:
                    last_error = Exception(f"Model {model_name} không tồn tại")
                    continue
                resp.raise_for_status()
                got_text = False
                buf = ""
                for raw in resp.iter_lines(decode_unicode=True):
                    if not raw or not raw.startswith("data:"):
                        continue
                    data_str = raw[5:].strip()
                    if data_str == "[DONE]":
                        break
                    try:
                        chunk = json.loads(data_str)
                        text = chunk["candidates"][0]["content"]["parts"][0]["text"]
                        if text:
                            got_text = True
                            yield text
                    except Exception:
                        continue
                if got_text:
                    return
            except ValueError:
                raise  # key/rate limit — raise ngay
            except Exception as e:
                last_error = e
                continue

    # Fallback OpenRouter (không stream)
    if or_key:
        try:
            history_text = "\n".join(
                f"{m['role'].upper()}: {m['parts'][0]['text']}" for m in messages_for_api
            )
            result = _call_openrouter(history_text, or_key)
            if result and not result.startswith("__OR_FAIL__"):
                yield result
                return
        except Exception as e:
            last_error = e

    raise RuntimeError(
        f"Tất cả model thất bại. Lỗi: {last_error}" if last_error
        else "Không kết nối được AI. Kiểm tra API key và mạng."
    )

def _build_data_context():
    """Tạo context tóm tắt dữ liệu hiện tại để đưa vào system prompt chat."""
    sheets=st.session_state.get("sheets",{})
    if not sheets: return "No data uploaded yet."
    parts=[]
    for name,df in sheets.items():
        stack=st.session_state["prep_transforms"].get(name,[])
        udf=_json_to_df(stack[-1][1]) if stack else df
        parts.append(df_summary(udf,name))
    results=st.session_state.get("run_results",[])
    results_txt=""
    if results:
        results_txt="\n\nML RESULTS AVAILABLE:\n"+pd.DataFrame(results).to_string(index=False)
    return "\n\n".join(parts)+results_txt

# ═══════════════════════════════════════════════════════════════════════════════
# DATA PREP
# ═══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def detect_problems(df_json):
    df=_json_to_df(df_json); n=len(df); problems=[]
    for col in df.columns:
        pct=df[col].isnull().mean()
        if pct>0:
            dtype="numeric" if pd.api.types.is_numeric_dtype(df[col]) else "categorical"
            problems.append({"col":col,"type":"missing","sev":"err" if pct>.3 else "warn",
                "msg":f'Column "{col}" — {pct:.1%} missing ({int(pct*n)} rows)',
                "fix":f"Fill with {'mean' if dtype=='numeric' else 'mode'}","dtype":dtype})
    for col in df.select_dtypes(include=[np.number]).columns:
        q1,q3=df[col].quantile(.25),df[col].quantile(.75); iqr=q3-q1
        if iqr>0:
            n_out=((df[col]<q1-3*iqr)|(df[col]>q3+3*iqr)).sum()
            if n_out>0:
                problems.append({"col":col,"type":"outlier","sev":"warn",
                    "msg":f'Column "{col}" — {n_out} extreme outliers ({n_out/n:.1%})',
                    "fix":"Cap to 3×IQR (Winsorize)","dtype":"numeric"})
    for col in df.select_dtypes(include=["object","string"]).columns:
        problems.append({"col":col,"type":"encoding","sev":"info",
            "msg":f'Column "{col}" is text — needs encoding for ML',"fix":"Label Encoding","dtype":"cat"})
    dups=df.duplicated().sum()
    if dups>0:
        problems.append({"col":"ALL","type":"duplicate","sev":"warn",
            "msg":f"{dups} duplicate rows ({dups/n:.1%})","fix":"Remove duplicates","dtype":"row"})
    num_cols=df.select_dtypes(include=[np.number]).columns.tolist()
    if len(num_cols)>1:
        rng=df[num_cols].max()-df[num_cols].min()
        if rng.max()>0 and (rng.max()/(rng.min()+1e-9))>100:
            problems.append({"col":"NUMERIC","type":"scale","sev":"info",
                "msg":"Numeric columns have very different scales — scaling recommended",
                "fix":"StandardScaler (auto-applied at training)","dtype":"numeric"})
    if not problems:
        problems.append({"col":"","type":"ok","sev":"ok",
            "msg":"No major data problems detected — dataset looks clean!","fix":"","dtype":""})
    return problems

@st.cache_data(show_spinner=False)
def fix_missing(df_json):
    df=_json_to_df(df_json); before=df.isnull().sum().sum()
    for col in df.columns:
        if df[col].isnull().any():
            if pd.api.types.is_numeric_dtype(df[col]): df[col]=df[col].fillna(df[col].mean())
            else:
                m=df[col].mode(); df[col]=df[col].fillna(m[0] if len(m) else "Unknown")
    return _df_to_json(df),f"Fixed {before-df.isnull().sum().sum()} missing values."

@st.cache_data(show_spinner=False)
def fix_duplicates(df_json):
    df=_json_to_df(df_json); before=len(df); df=df.drop_duplicates()
    return _df_to_json(df),f"Removed {before-len(df)} duplicates. {len(df):,} rows remain."

@st.cache_data(show_spinner=False)
def fix_outliers(df_json):
    """Cap outlier: giới hạn giá trị tại ngưỡng 3×IQR, không xóa dòng."""
    df=_json_to_df(df_json); capped=0
    for col in df.select_dtypes(include=[np.number]).columns:
        q1,q3=df[col].quantile(.25),df[col].quantile(.75); iqr=q3-q1
        if iqr>0:
            lo,hi=q1-3*iqr,q3+3*iqr
            capped+=((df[col]<lo)|(df[col]>hi)).sum(); df[col]=df[col].clip(lower=lo,upper=hi)
    return _df_to_json(df),f"Capped {capped} extreme outlier values (3×IQR) — dòng vẫn được giữ lại, chỉ đổi giá trị."

def remove_outliers(df_json):
    """Clean outlier: xóa hẳn các dòng có ít nhất 1 cột vượt ngưỡng 3×IQR."""
    df=_json_to_df(df_json)
    before=len(df); mask=pd.Series([True]*before,index=df.index)
    for col in df.select_dtypes(include=[np.number]).columns:
        q1,q3=df[col].quantile(.25),df[col].quantile(.75); iqr=q3-q1
        if iqr>0:
            lo,hi=q1-3*iqr,q3+3*iqr
            mask=mask&(df[col]>=lo)&(df[col]<=hi)
    df=df[mask].reset_index(drop=True)
    removed=before-len(df)
    return _df_to_json(df),f"Đã xóa {removed} dòng chứa outlier ({removed/before:.1%}). Còn lại {len(df):,} dòng."

@st.cache_data(show_spinner=False)
def fix_encode(df_json):
    df=_json_to_df(df_json)
    text_cols=df.select_dtypes(include=["object","string"]).columns.tolist()
    if not text_cols: return _df_to_json(df),"No text columns to encode.",{}
    mapping={}; le=LabelEncoder()
    for col in text_cols:
        le.fit(df[col].astype(str))
        mapping[col]={str(c):int(i) for i,c in enumerate(le.classes_)}
        df[col]=le.transform(df[col].astype(str))
    return _df_to_json(df),f"Label-encoded {len(text_cols)} column(s): {', '.join(text_cols)}.",mapping

def suggest_target(df):
    kws=["label","target","class","churn","default","outcome","result","status",
         "output","dependent","response","predict","category","nhãn"]
    for col in df.columns:
        if any(k in col.lower().replace(" ","_") for k in kws): return col
    last=df.columns[-1]
    return last if df[last].nunique()<=20 else None

# ═══════════════════════════════════════════════════════════════════════════════
# ML RUNNERS
# ═══════════════════════════════════════════════════════════════════════════════
def _dark_fig(w=6,h=4):
    fig,ax=plt.subplots(figsize=(w,h))
    fig.patch.set_facecolor('#0d1117'); ax.set_facecolor('#111827')
    ax.tick_params(colors='#9ca3af',labelcolor='#9ca3af')
    for sp in ax.spines.values(): sp.set_edgecolor('#374151')
    return fig,ax

def run_classification(method,df,target,features,test_size,balance):
    df_enc=encode_df(df[features+[target]].dropna())
    scaler=StandardScaler()
    X=scaler.fit_transform(df_enc[features].values); y=df_enc[target].values
    if balance=="Random Oversampling":
        X,y=RandomOverSampler(random_state=42).fit_resample(X,y); st.info("✅ Applied Random Oversampling.")
    elif balance=="SMOTE":
        try: X,y=SMOTE(random_state=42).fit_resample(X,y); st.info("✅ Applied SMOTE.")
        except Exception as e: st.warning(f"SMOTE skipped: {e}")
    X_tr,X_te,y_tr,y_te=train_test_split(X,y,test_size=test_size,random_state=42)
    models={
        "Logistic Regression":LogisticRegression(max_iter=1000),
        "Linear Discriminant Analysis (LDA)":LinearDiscriminantAnalysis(),
        "K-Nearest Neighbors (KNN)":KNeighborsClassifier(),
        "Classification Trees":DecisionTreeClassifier(max_depth=5),
        "Naive Bayes":GaussianNB(),
        "Support Vector Machine (SVM)":SVC(probability=True),
        "Random Forest":RandomForestClassifier(n_estimators=100,random_state=42),
        "Neural Networks (MLP)":MLPClassifier(max_iter=500,random_state=42),
    }
    mdl=models[method]; mdl.fit(X_tr,y_tr); y_pred=mdl.predict(X_te)
    acc=accuracy_score(y_te,y_pred); is_bin=len(np.unique(y))==2; avg="binary" if is_bin else "weighted"
    prec=precision_score(y_te,y_pred,average=avg,zero_division=0)
    rec=recall_score(y_te,y_pred,average=avg,zero_division=0)
    f1=f1_score(y_te,y_pred,average=avg,zero_division=0)
    try:
        auc=roc_auc_score(y_te,mdl.predict_proba(X_te)[:,1] if is_bin else mdl.predict_proba(X_te),
            multi_class="ovr" if not is_bin else "raise",average="weighted" if not is_bin else None)
    except Exception: auc=None
    metrics={"Method":method,"Sheet":st.session_state.get("active_sheet",""),
             "Accuracy":f"{acc:.4f}","Precision":f"{prec:.4f}","Recall":f"{rec:.4f}",
             "F1-Score":f"{f1:.4f}","AUC":f"{auc:.4f}" if auc else "N/A",
             "Train rows":len(X_tr),"Validation rows":len(X_te)}
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Accuracy",f"{acc:.2%}"); c2.metric("F1",f"{f1:.4f}")
    c3.metric("AUC",f"{auc:.4f}" if auc else "N/A"); c4.metric("Precision",f"{prec:.4f}")
    with st.expander("📋 Full classification report",expanded=False):
        st.text(classification_report(y_te,y_pred))
    col_a,col_b=st.columns(2)
    with col_a:
        st.markdown('<div class="chart-title">Confusion Matrix</div>',unsafe_allow_html=True)
        fig,ax=_dark_fig(5,4)
        cm = confusion_matrix(y_te, y_pred)
        sns.heatmap(cm, annot=False, fmt='d', cmap='Blues', ax=ax,
                    linewidths=.5, linecolor='#374151')
        # Chữ đen trên tất cả ô — dễ đọc nhất
        for _i in range(cm.shape[0]):
            for _j in range(cm.shape[1]):
                ax.text(_j + 0.5, _i + 0.5, str(cm[_i, _j]),
                        ha='center', va='center', fontsize=14, fontweight='bold', color='black')
        ax.set_title("Confusion Matrix",color='#e5e7eb')
        ax.set_xlabel("Predicted",color='#9ca3af'); ax.set_ylabel("Actual",color='#9ca3af')
        fig_to_st(fig, save_key=method)
        st.markdown('<div class="chart-cap">Large diagonal = correct predictions. Off-diagonal = errors.</div>',unsafe_allow_html=True)
    with col_b:
        if is_bin and hasattr(mdl,"predict_proba"):
            st.markdown('<div class="chart-title">ROC Curve</div>',unsafe_allow_html=True)
            try:
                fpr,tpr,_=roc_curve(y_te,mdl.predict_proba(X_te)[:,1])
                fig2,ax2=_dark_fig(5,4)
                ax2.plot(fpr,tpr,color='#3b82f6',lw=2,label=f"AUC={auc:.3f}" if auc else "ROC")
                ax2.plot([0,1],[0,1],'r--',lw=1,label="Random")
                ax2.set_xlabel("FPR",color='#9ca3af'); ax2.set_ylabel("TPR",color='#9ca3af')
                ax2.set_title("ROC Curve",color='#e5e7eb')
                ax2.legend(facecolor='#111827',labelcolor='#9ca3af')
                fig_to_st(fig2, save_key=method)
                st.markdown('<div class="chart-cap">Curve near top-left = strong model. Near diagonal = weak.</div>',unsafe_allow_html=True)
            except Exception: pass
    if hasattr(mdl,"feature_importances_") or hasattr(mdl,"coef_"):
        st.markdown('<div class="chart-title">Feature Importance</div>',unsafe_allow_html=True)
        if hasattr(mdl,"feature_importances_"):
            fi=pd.Series(mdl.feature_importances_,index=features).sort_values(ascending=False)
        else:
            fi=pd.Series(np.abs(mdl.coef_[0] if mdl.coef_.ndim>1 else mdl.coef_),index=features).sort_values(ascending=False)
        _top_n=min(len(fi),20)
        fig3,ax3=_dark_fig(7,max(3,min(14,_top_n*.4)))
        fi.head(_top_n).plot(kind='barh',ax=ax3,color='#3b82f6')
        ax3.set_title(f"Top {_top_n} Important Features",color='#e5e7eb'); ax3.invert_yaxis()
        plt.tight_layout(); fig_to_st(fig3, save_key=method)
        # Hiển thị bảng feature importance
        fi_df=fi.head(_top_n).reset_index()
        fi_df.columns=["Feature","Importance"]
        fi_df["Importance"]=fi_df["Importance"].round(4)
        st.session_state.setdefault("_fi_tables",{})[method]=fi_df
    if method=="Classification Trees":
        with st.expander("🌿 Decision Tree Rules",expanded=False):
            st.code(export_text(mdl,feature_names=features,max_depth=4),language="")

    # Lưu model + metadata để dùng cho Predict New Data
    st.session_state.setdefault("_trained_models",{})[method]={
        "model":mdl,"scaler":scaler,"features":features,"target":target,
        "type":"classification","classes":list(np.unique(y)),
        "label_encoder": st.session_state.get("enc_mapping",{}).get(st.session_state.get("active_sheet",""),{})
    }
    # Lưu train/val data
    tr_df=pd.DataFrame(X_tr,columns=[f"{c}_scaled" for c in features])
    tr_df[target]=y_tr
    te_df=pd.DataFrame(X_te,columns=[f"{c}_scaled" for c in features])
    te_df[target+"_actual"]=y_te; te_df[target+"_predicted"]=y_pred
    st.session_state.setdefault("_split_data",{})[method]={
        "train":tr_df,"validation":te_df,"features":features,"target":target
    }
    return metrics,mdl,X_te,y_te,y_pred

def run_regression(method,df,target,features,test_size):
    df_enc=encode_df(df[features+[target]].dropna())
    scaler_r=StandardScaler()
    X=scaler_r.fit_transform(df_enc[features].values); y=df_enc[target].values
    X_tr,X_te,y_tr,y_te=train_test_split(X,y,test_size=test_size,random_state=42)
    mdl=LinearRegression() if method=="Linear Regression" else MLPRegressor(max_iter=500,random_state=42)
    mdl.fit(X_tr,y_tr); y_pred=mdl.predict(X_te)
    mse=mean_squared_error(y_te,y_pred); r2=r2_score(y_te,y_pred); residuals=y_te-y_pred
    metrics={"Method":method,"Sheet":st.session_state.get("active_sheet",""),
             "R²":f"{r2:.4f}","RMSE":f"{np.sqrt(mse):.4f}","Train rows":len(X_tr),"Validation rows":len(X_te)}
    c1,c2=st.columns(2); c1.metric("R² Score",f"{r2:.4f}"); c2.metric("RMSE",f"{np.sqrt(mse):.4f}")
    if r2>=.7: st.success(f"R²={r2:.4f} — Good fit ({r2:.1%} variance explained).")
    elif r2>=.4: st.warning(f"R²={r2:.4f} — Moderate. Consider adding features.")
    else: st.error(f"R²={r2:.4f} — Weak fit.")
    col_a,col_b=st.columns(2)
    with col_a:
        st.markdown('<div class="chart-title">Actual vs Predicted</div>',unsafe_allow_html=True)
        fig,ax=_dark_fig(5,4)
        ax.scatter(y_te,y_pred,alpha=.5,color='#3b82f6',s=20)
        mn,mx=min(y_te.min(),y_pred.min()),max(y_te.max(),y_pred.max())
        ax.plot([mn,mx],[mn,mx],'r--',lw=1.5)
        ax.set_xlabel("Actual",color='#9ca3af'); ax.set_ylabel("Predicted",color='#9ca3af')
        ax.set_title("Actual vs Predicted",color='#e5e7eb'); fig_to_st(fig, save_key=method)
    with col_b:
        st.markdown('<div class="chart-title">Residual Plot</div>',unsafe_allow_html=True)
        fig2,ax2=_dark_fig(5,4)
        ax2.scatter(y_pred,residuals,alpha=.5,color='#a78bfa',s=20)
        ax2.axhline(0,color='#ef4444',linestyle='--',lw=1.5)
        ax2.set_xlabel("Predicted",color='#9ca3af'); ax2.set_ylabel("Residual",color='#9ca3af')
        ax2.set_title("Residual Plot",color='#e5e7eb'); fig_to_st(fig2)
    # Lưu model + split data
    st.session_state.setdefault("_trained_models",{})[method]={
        "model":mdl,"scaler":scaler_r,"features":features,"target":target,"type":"regression"
    }
    tr_df2=pd.DataFrame(X_tr,columns=[f"{c}_scaled" for c in features]); tr_df2[target]=y_tr
    te_df2=pd.DataFrame(X_te,columns=[f"{c}_scaled" for c in features])
    te_df2[target+"_actual"]=y_te; te_df2[target+"_predicted"]=y_pred
    st.session_state.setdefault("_split_data",{})[method]={"train":tr_df2,"validation":te_df2,"features":features,"target":target}
    return metrics

def run_clustering(method,df,features,n_clusters):
    df_enc=encode_df(df[features].dropna()); X=StandardScaler().fit_transform(df_enc.values)
    if method=="K-Means Clustering":
        mdl=KMeans(n_clusters=n_clusters,random_state=42,n_init=10); labels=mdl.fit_predict(X)
        k_range=range(2,min(11,len(X))); inertias=[]
        for k in k_range:
            km=KMeans(n_clusters=k,random_state=42,n_init=10); km.fit(X); inertias.append(km.inertia_)
        st.markdown('<div class="chart-title">Elbow Curve</div>',unsafe_allow_html=True)
        fig,ax=_dark_fig(6,3)
        ax.plot(list(k_range),inertias,'o-',color='#3b82f6',lw=2)
        ax.axvline(n_clusters,color='#f472b6',linestyle='--',lw=1.5,label=f"K={n_clusters}")
        ax.set_title("Elbow Curve",color='#e5e7eb'); ax.set_xlabel("K",color='#9ca3af'); ax.set_ylabel("Inertia",color='#9ca3af')
        ax.legend(facecolor='#111827',labelcolor='#9ca3af'); fig_to_st(fig)
    else:
        mdl=AgglomerativeClustering(n_clusters=n_clusters); labels=mdl.fit_predict(X)
        st.markdown('<div class="chart-title">Dendrogram</div>',unsafe_allow_html=True)
        linked=linkage(X[:min(200,len(X))],method='ward')
        fig,ax=_dark_fig(8,4)
        dendrogram(linked,ax=ax,color_threshold=0,above_threshold_color='#3b82f6',leaf_font_size=6)
        ax.set_title("Dendrogram (sample 200)",color='#e5e7eb'); plt.tight_layout(); fig_to_st(fig)
    try:
        sil=silhouette_score(X,labels); st.metric("Silhouette Score",f"{sil:.4f}")
        if sil>.7: st.success(f"Excellent separation (Silhouette={sil:.4f})")
        elif sil>.5: st.info(f"Good separation (Silhouette={sil:.4f})")
        elif sil>.25: st.warning("Moderate separation — try different K")
        else: st.error("Weak clusters — consider different K or features")
    except Exception: sil=None
    st.markdown('<div class="chart-title">Cluster Scatter (PCA 2D)</div>',unsafe_allow_html=True)
    n_comp=min(2,X.shape[1]); pca=PCA(n_components=n_comp,random_state=42); X_2d=pca.fit_transform(X)
    var_exp=pca.explained_variance_ratio_.sum()*100
    fig2,ax2=_dark_fig(7,5); pal=plt.cm.tab10.colors
    for c in np.unique(labels):
        mask=labels==c
        ax2.scatter(X_2d[mask,0],X_2d[mask,1] if n_comp>1 else np.zeros(mask.sum()),
                    color=pal[c%10],label=f"Cluster {c}",alpha=.7,s=25)
    ax2.legend(facecolor='#111827',labelcolor='#9ca3af',fontsize=8)
    ax2.set_title(f"PCA 2D ({var_exp:.1f}% variance)",color='#e5e7eb')
    ax2.set_xlabel("PC1",color='#9ca3af'); ax2.set_ylabel("PC2",color='#9ca3af')
    plt.tight_layout(); fig_to_st(fig2)
    df_out=df[features].copy(); df_out["Cluster"]=labels
    metrics={"Method":method,"Sheet":st.session_state.get("active_sheet",""),
             "K":n_clusters,"Silhouette":f"{sil:.4f}" if sil else "N/A","Rows":len(df_out)}
    return metrics,df_out

def run_association(df,min_support,min_confidence,min_lift):
    records=[]
    for _,row in df.iterrows():
        items=[str(v).strip() for v in row.dropna().values if str(v).strip()]
        if items: records.append(items)
    if not records: st.error("Could not parse transactional data."); return None,None
    te=TransactionEncoder(); te_arr=te.fit_transform(records)
    df_bool=pd.DataFrame(te_arr,columns=te.columns_)
    freq=apriori(df_bool,min_support=min_support,use_colnames=True)
    if freq.empty: st.warning("No frequent itemsets. Lower min support."); return None,None
    rules=association_rules(freq,metric="lift",min_threshold=min_lift)
    rules=rules[rules["confidence"]>=min_confidence].sort_values("lift",ascending=False)
    st.success(f"Found {len(rules)} rules from {len(freq)} frequent itemsets.")
    display=rules[["antecedents","consequents","support","confidence","lift"]].head(20).copy()
    display["antecedents"]=display["antecedents"].apply(lambda x:", ".join(list(x)))
    display["consequents"]=display["consequents"].apply(lambda x:", ".join(list(x)))
    st.dataframe(_sanitize_df(display),width='stretch')
    if not rules.empty:
        fig,ax=_dark_fig(7,4)
        sc=ax.scatter(rules["support"],rules["confidence"],c=rules["lift"],cmap="plasma",alpha=.8,s=60)
        plt.colorbar(sc,ax=ax,label="Lift")
        ax.set_xlabel("Support",color='#9ca3af'); ax.set_ylabel("Confidence",color='#9ca3af')
        ax.set_title("Support vs Confidence (color=Lift)",color='#e5e7eb'); fig_to_st(fig)
    metrics={"Method":"Association Rules","Sheet":st.session_state.get("active_sheet",""),
             "Rules found":len(rules),"Min Support":min_support,"Min Confidence":min_confidence,"Min Lift":min_lift}
    return metrics,display

def export_to_excel(results_dict, figures_dict=None):
    """
    Xuất Excel với dữ liệu + biểu đồ nhúng vào cùng sheet.
    figures_dict: {method_name: [png_bytes, ...]}
    """
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in results_dict.items():
            df.to_excel(writer, sheet_name=str(name)[:31], index=False)

        # Nhúng biểu đồ vào sheet riêng "Charts"
        if figures_dict:
            # Tạo sheet trắng "Charts"
            wb = writer.book
            if "Charts" not in wb.sheetnames:
                ws_charts = wb.create_sheet("Charts")
            else:
                ws_charts = wb["Charts"]

            row_offset = 1
            for method_name, fig_bytes_list in figures_dict.items():
                chart_labels = ["Confusion_Matrix","ROC_Curve","Feature_Importance",
                                "Actual_vs_Predicted","Residuals","Elbow_Chart","Cluster_Plot"]
                # Ghi tên model
                ws_charts.cell(row=row_offset, column=1, value=f"▶ {method_name}")
                ws_charts.cell(row=row_offset, column=1).font = \
                    __import__('openpyxl').styles.Font(bold=True, size=12, color="3B82F6")
                row_offset += 1

                col_offset = 1
                for i, img_bytes in enumerate(fig_bytes_list):
                    try:
                        img_io = io.BytesIO(img_bytes)
                        xl_img = XLImage(img_io)
                        # Scale xuống vừa vặn (khoảng 400x280 pixels)
                        xl_img.width = 380
                        xl_img.height = 260
                        # Đặt ảnh: mỗi ảnh cách nhau 7 cột (~ 400px)
                        col_letter = chr(ord('A') + (col_offset - 1) * 7)
                        cell_ref = f"{col_letter}{row_offset}"
                        ws_charts.add_image(xl_img, cell_ref)
                        # Ghi label bên dưới ảnh
                        label = chart_labels[i] if i < len(chart_labels) else f"Chart_{i+1}"
                        ws_charts.cell(
                            row=row_offset + 15,
                            column=(col_offset - 1) * 7 + 1,
                            value=label.replace("_", " ")
                        )
                        col_offset += 1
                    except Exception:
                        pass
                # Mỗi model chiếm ~18 dòng
                row_offset += 19

    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# EDA PROFILING
# ═══════════════════════════════════════════════════════════════════════════════
def render_eda_section(df):
    num_cols=df.select_dtypes(include=[np.number]).columns.tolist()
    cat_cols=df.select_dtypes(include=["object","string"]).columns.tolist()
    tab_ov,tab_dist,tab_corr,tab_miss=st.tabs(["📋 Overview","📊 Distributions","🔗 Correlations","⚠️ Missing & Outliers"])

    with tab_ov:
        profile=[]
        for col in df.columns:
            miss=df[col].isnull().sum(); miss_p=f"{df[col].isnull().mean():.1%}"; uniq=df[col].nunique()
            sample=str(df[col].dropna().iloc[0]) if df[col].dropna().shape[0]>0 else ""
            if pd.api.types.is_numeric_dtype(df[col]):
                stats=f"min={df[col].min():.3g} / mean={df[col].mean():.3g} / max={df[col].max():.3g}"
            else:
                stats=", ".join([str(v) for v in df[col].value_counts().head(3).index.tolist()])
            profile.append({"Column":col,"Type":str(df[col].dtype),"Missing":miss,"Missing%":miss_p,
                            "Unique":uniq,"Stats / Top Values":stats,"Sample":sample[:40]})
        st.dataframe(_sanitize_df(pd.DataFrame(profile)),width='stretch',height=min(600,40*len(df.columns)+50))

    with tab_dist:
        if not num_cols: st.info("No numeric columns for distribution charts.")
        else:
            cols_p=num_cols[:12]; n_c=3; n_r=int(np.ceil(len(cols_p)/n_c))
            fig,axes=plt.subplots(n_r,n_c,figsize=(14,n_r*3)); fig.patch.set_facecolor('#0d1117')
            axf=axes.flatten() if n_r*n_c>1 else [axes]
            for i,col in enumerate(cols_p):
                ax=axf[i]; ax.set_facecolor('#111827')
                for sp in ax.spines.values(): sp.set_edgecolor('#374151')
                ax.tick_params(colors='#6b7280',labelsize=7)
                data=df[col].dropna(); n_bins=min(30,max(10,int(len(data)**0.5)))
                ax.hist(data,bins=n_bins,color='#3b82f6',alpha=.75,edgecolor='#1e40af')
                ax.axvline(data.mean(),color='#f472b6',lw=1.5,linestyle='--')
                ax.set_title(col[:22],color='#e5e7eb',fontsize=8,fontweight='600')
                ax.set_ylabel("Count",color='#6b7280',fontsize=7)
            for j in range(len(cols_p),len(axf)): axf[j].set_visible(False)
            plt.tight_layout(pad=1.5); fig_to_st(fig)
            st.markdown('<div class="chart-cap">Pink line = column mean. Shows value distribution per numeric column.</div>',unsafe_allow_html=True)
        if cat_cols:
            st.markdown("---")
            st.markdown('<div class="chart-title">Categorical Frequencies</div>',unsafe_allow_html=True)
            cat_p=[c for c in cat_cols if df[c].nunique()<=30][:6]
            if cat_p:
                n_c2=min(3,len(cat_p)); n_r2=int(np.ceil(len(cat_p)/n_c2))
                fig2,axes2=plt.subplots(n_r2,n_c2,figsize=(14,n_r2*3.2)); fig2.patch.set_facecolor('#0d1117')
                axf2=axes2.flatten() if n_r2*n_c2>1 else [axes2]
                for i,col in enumerate(cat_p):
                    ax2=axf2[i]; ax2.set_facecolor('#111827')
                    for sp in ax2.spines.values(): sp.set_edgecolor('#374151')
                    vc=df[col].value_counts().head(15); colors=plt.cm.Set2(np.linspace(0,.8,len(vc)))
                    ax2.barh(range(len(vc)),vc.values,color=colors,alpha=.85)
                    ax2.set_yticks(range(len(vc)))
                    ax2.set_yticklabels([str(v)[:20] for v in vc.index],fontsize=7,color='#9ca3af')
                    ax2.tick_params(colors='#6b7280',labelsize=7)
                    ax2.set_title(col[:22],color='#e5e7eb',fontsize=8,fontweight='600'); ax2.invert_yaxis()
                for j in range(len(cat_p),len(axf2)): axf2[j].set_visible(False)
                plt.tight_layout(pad=1.5); fig_to_st(fig2)

    with tab_corr:
        if len(num_cols)<2: st.info("Need at least 2 numeric columns for correlation analysis.")
        else:
            corr=df[num_cols].corr(); sz=max(6,min(14,len(num_cols)*.6))
            fig,ax=plt.subplots(figsize=(sz,sz*.85)); fig.patch.set_facecolor('#0d1117'); ax.set_facecolor('#111827')
            mask=np.zeros_like(corr,dtype=bool); mask[np.triu_indices_from(mask)]=True
            sns.heatmap(corr,mask=mask,annot=len(num_cols)<=15,fmt=".2f",cmap='RdBu_r',center=0,ax=ax,
                        linewidths=.3,linecolor='#1f2937',annot_kws={"size":7,"color":"white"},
                        cbar_kws={"shrink":.6})
            ax.set_title("Feature Correlation Matrix (lower triangle)",color='#e5e7eb',pad=10)
            ax.tick_params(colors='#9ca3af',labelsize=8,rotation=45)
            plt.tight_layout(); fig_to_st(fig)
            st.markdown('<div class="chart-cap">Red=positive · Blue=negative · White=no correlation. Values >0.8 may indicate redundant features.</div>',unsafe_allow_html=True)
            strong=[]
            for i in range(len(corr)):
                for j in range(i+1,len(corr)):
                    val=corr.iloc[i,j]
                    if abs(val)>=.75:
                        strong.append({"Feature A":corr.columns[i],"Feature B":corr.columns[j],
                                       "Correlation":f"{val:.3f}","Strength":"🔴 Strong (≥0.9)" if abs(val)>=.9 else "🟡 Moderate-Strong"})
            if strong:
                st.markdown("**⚡ Highly correlated pairs (|r| ≥ 0.75):**")
                st.dataframe(_sanitize_df(pd.DataFrame(strong)),width='stretch')

    with tab_miss:
        miss=df.isnull().sum(); miss=miss[miss>0].sort_values(ascending=False)
        if miss.empty: st.success("✅ No missing values found in this dataset.")
        else:
            fig,ax=_dark_fig(9,max(3,len(miss)*.4)); pcts=(miss/len(df)*100)
            colors=['#ef4444' if p>30 else '#f59e0b' if p>10 else '#3b82f6' for p in pcts]
            ax.barh(range(len(miss)),pcts.values,color=colors,alpha=.85)
            ax.set_yticks(range(len(miss))); ax.set_yticklabels(miss.index.tolist(),color='#9ca3af',fontsize=8)
            ax.set_xlabel("Missing %",color='#9ca3af'); ax.set_title("Missing Values by Column",color='#e5e7eb')
            ax.axvline(30,color='#ef4444',linestyle='--',lw=1,alpha=.5,label="30% threshold")
            ax.legend(facecolor='#111827',labelcolor='#9ca3af',fontsize=8)
            ax.invert_yaxis(); plt.tight_layout(); fig_to_st(fig)
            st.markdown('<div class="chart-cap">Red=>30% critical · Amber=10-30% moderate · Blue=<10% manageable.</div>',unsafe_allow_html=True)
        if num_cols:
            st.markdown("---")
            st.markdown('<div class="chart-title">Outlier Box Plots</div>',unsafe_allow_html=True)
            box_cols=num_cols[:10]; n_c3=min(4,len(box_cols)); n_r3=int(np.ceil(len(box_cols)/n_c3))
            fig2,axes2=plt.subplots(n_r3,n_c3,figsize=(13,n_r3*3)); fig2.patch.set_facecolor('#0d1117')
            axf3=axes2.flatten() if n_r3*n_c3>1 else [axes2]
            for i,col in enumerate(box_cols):
                ax3=axf3[i]; ax3.set_facecolor('#111827')
                for sp in ax3.spines.values(): sp.set_edgecolor('#374151')
                ax3.boxplot(df[col].dropna(),patch_artist=True,
                    boxprops=dict(facecolor='#1d4ed8',color='#60a5fa'),
                    whiskerprops=dict(color='#60a5fa'),capprops=dict(color='#60a5fa'),
                    medianprops=dict(color='#f472b6',lw=2),
                    flierprops=dict(markerfacecolor='#ef4444',markersize=4,marker='o'))
                ax3.set_title(col[:18],color='#e5e7eb',fontsize=8,fontweight='600')
                ax3.tick_params(colors='#6b7280',labelsize=7)
            for j in range(len(box_cols),len(axf3)): axf3[j].set_visible(False)
            plt.tight_layout(pad=1.5); fig_to_st(fig2)
            st.markdown('<div class="chart-cap">Pink=median · Dots beyond whiskers=outliers · Wide spread=high variance.</div>',unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SMART ALGORITHM ADVISOR
# ═══════════════════════════════════════════════════════════════════════════════
ALGO_CATEGORIES = {
    "supervised_clf": {
        "title":"Supervised Learning — Classification",
        "icon":"🎯","color":"#60a5fa","bg":"rgba(59,130,246,.08)","border":"rgba(59,130,246,.25)",
        "desc":"Bạn có target variable dạng category. Model học từ dữ liệu có nhãn để phân loại.",
        "when":"Dự đoán: churn/no-churn, fraud/normal, loại sản phẩm, risk level, etc.",
    },
    "supervised_reg": {
        "title":"Supervised Learning — Regression",
        "icon":"📈","color":"#a78bfa","bg":"rgba(124,58,237,.08)","border":"rgba(124,58,237,.25)",
        "desc":"Bạn có target variable dạng số liên tục. Model học để dự báo giá trị.",
        "when":"Dự báo: giá, doanh thu, nhiệt độ, số lượng, chi phí, etc.",
    },
    "unsupervised_clus": {
        "title":"Unsupervised Learning — Clustering",
        "icon":"🔮","color":"#34d399","bg":"rgba(16,185,129,.08)","border":"rgba(16,185,129,.25)",
        "desc":"Không có target. Model tự tìm nhóm tự nhiên ẩn trong dữ liệu.",
        "when":"Phân khúc khách hàng, phân nhóm sản phẩm, anomaly detection.",
    },
    "unsupervised_assoc": {
        "title":"Unsupervised Learning — Association Rules",
        "icon":"🔗","color":"#f472b6","bg":"rgba(244,114,182,.08)","border":"rgba(244,114,182,.25)",
        "desc":"Tìm pattern đồng xuất hiện trong dữ liệu giao dịch.",
        "when":"Market basket, cross-sell, maintenance co-occurrence, log pattern.",
    },
}

# Rule table: (goal_type, size, interpretability) → [(method, tier, why)]
_R = {
    ("clf","small","high"):[
        ("Classification Trees","primary","Dataset nhỏ + cần giải thích → cây quyết định IF/THEN tốt nhất"),
        ("Logistic Regression","alt","Hệ số dễ diễn giải, phù hợp report cho management"),
        ("Naive Bayes","opt","Rất nhanh, tốt làm baseline cho tập nhỏ"),
    ],
    ("clf","small","low"):[
        ("Random Forest","primary","Ensemble mạnh dù tập nhỏ — tránh overfit tốt hơn single tree"),
        ("K-Nearest Neighbors (KNN)","alt","Non-parametric, không cần giả định phân phối"),
        ("Classification Trees","opt","Backup khi cần diễn giải"),
    ],
    ("clf","medium","high"):[
        ("Classification Trees","primary","Dễ trình bày IF/THEN cho business stakeholders"),
        ("Logistic Regression","alt","Coefficient = tác động trực tiếp của từng feature"),
        ("Linear Discriminant Analysis (LDA)","opt","Tốt khi các class có phân phối normal"),
    ],
    ("clf","medium","low"):[
        ("Random Forest","primary","Best overall cho tập trung bình — robust & accurate"),
        ("Support Vector Machine (SVM)","alt","Mạnh với nhiều features, tốt với margin rõ ràng"),
        ("Neural Networks (MLP)","opt","Nếu pattern phức tạp phi tuyến tính"),
    ],
    ("clf","large","high"):[
        ("Logistic Regression","primary","Scale tốt với tập lớn, hệ số rõ ràng"),
        ("Classification Trees","alt","Vẫn interpretable với tập lớn nếu giới hạn depth"),
        ("Linear Discriminant Analysis (LDA)","opt","Nếu features numeric và normal-distributed"),
    ],
    ("clf","large","low"):[
        ("Random Forest","primary","Best overall cho tập lớn — ổn định, chính xác"),
        ("Neural Networks (MLP)","alt","Khai thác pattern phi tuyến phức tạp"),
        ("Support Vector Machine (SVM)","opt","Nếu features high-dimensional"),
    ],
    ("reg","small","high"):[
        ("Linear Regression","primary","Interpretable nhất — coefficient = tác động trực tiếp"),
        ("Neural Networks Regression (MLP)","opt","Chỉ thử nếu linear không fit tốt (R²<0.5)"),
    ],
    ("reg","small","low"):[
        ("Linear Regression","primary","Baseline tốt, dễ validate với tập nhỏ"),
        ("Neural Networks Regression (MLP)","alt","Nếu relationship rõ ràng là phi tuyến"),
    ],
    ("reg","medium","high"):[
        ("Linear Regression","primary","Tốt nhất cho interpretability"),
        ("Neural Networks Regression (MLP)","alt","Thử nếu R² < 0.6 với linear"),
    ],
    ("reg","medium","low"):[
        ("Neural Networks Regression (MLP)","primary","Tốt hơn linear cho pattern phức tạp"),
        ("Linear Regression","alt","Baseline để so sánh"),
    ],
    ("reg","large","high"):[
        ("Linear Regression","primary","Scale tốt với tập lớn, dễ deploy"),
        ("Neural Networks Regression (MLP)","alt","Nếu có non-linearity rõ ràng"),
    ],
    ("reg","large","low"):[
        ("Neural Networks Regression (MLP)","primary","Tốt nhất cho tập lớn phi tuyến"),
        ("Linear Regression","alt","Benchmark cơ bản luôn cần chạy"),
    ],
    ("clus","any","any"):[
        ("K-Means Clustering","primary","Nhanh, scalable, dễ interpret — luôn bắt đầu đây"),
        ("Hierarchical Clustering","alt","Không cần chỉ định K trước, xem dendrogram để chọn"),
    ],
    ("assoc","any","any"):[
        ("Association Rules (Apriori)","primary","Tìm IF-THEN patterns trong dữ liệu giao dịch"),
    ],
}

def get_recommendations(df, goal_type, interp, is_imbalanced):
    n=len(df)
    size="small" if n<500 else ("medium" if n<10000 else "large")
    if goal_type=="clf":
        cat="supervised_clf"
        key=("clf",size,interp)
        recs=_R.get(key,_R[("clf","medium",interp)])
        prereqs=[]
        if is_imbalanced=="yes":
            prereqs=[
                ("SMOTE","prereq","⚠️ Class imbalance → chạy trước khi classification để tránh bias"),
                ("Random Oversampling","prereq","Alternative đơn giản hơn SMOTE cho tập nhỏ"),
            ]
        return cat, prereqs+recs, size, n
    elif goal_type=="reg":
        cat="supervised_reg"
        key=("reg",size,interp)
        recs=_R.get(key,_R[("reg","medium",interp)])
        return cat, recs, size, n
    elif goal_type=="clus":
        return "unsupervised_clus", _R[("clus","any","any")], size, n
    else:
        return "unsupervised_assoc", _R[("assoc","any","any")], size, n


def render_algo_advisor(df):
    st.markdown(
        '<div style="background:linear-gradient(135deg,#0f1f3d,#0d1120);border:1px solid #1e3a5f;'
        'border-radius:12px;padding:14px 18px;margin-bottom:14px;">'
        '<div style="font-size:.93rem;font-weight:700;color:#60a5fa;margin-bottom:3px">🎯 Smart Algorithm Advisor</div>'
        '<div style="font-size:.78rem;color:#6b7280">Trả lời 4 câu hỏi → hệ thống gợi ý thuật toán phù hợp nhất cho dữ liệu của bạn.</div>'
        '</div>', unsafe_allow_html=True)

    n_rows=len(df); num_cols=df.select_dtypes(include=[np.number]).columns.tolist()

    # Q1 — Goal
    st.markdown('<div class="quiz-card"><div class="quiz-q">1️⃣  Mục tiêu phân tích của bạn là gì?</div>'
                '<div class="quiz-hint">Bạn muốn dự đoán kết quả cụ thể hay khám phá cấu trúc ẩn trong dữ liệu?</div>',
                unsafe_allow_html=True)
    q1=st.radio("q1",["🎯  Dự đoán một kết quả (classification / regression)",
                       "🔮  Phân nhóm / tìm cụm tự nhiên (clustering)",
                       "🔗  Tìm items thường đi cùng nhau (association rules)"],
                key="adv_q1",label_visibility="collapsed")
    st.markdown('</div>',unsafe_allow_html=True)
    goal_type="clf" if "Dự đoán" in q1 else ("clus" if "Phân nhóm" in q1 else "assoc")

    # Q2 — Target type (supervised only)
    if goal_type=="clf":
        st.markdown('<div class="quiz-card"><div class="quiz-q">2️⃣  Target variable (kết quả cần dự đoán) thuộc loại nào?</div>',
                    unsafe_allow_html=True)
        q2=st.radio("q2",["📌  Category / nhãn rời rạc (Yes/No, Loại A/B/C, Churn/No-Churn)",
                           "📊  Số liên tục (giá tiền, doanh thu, nhiệt độ, số lượng)"],
                    key="adv_q2",label_visibility="collapsed")
        st.markdown('</div>',unsafe_allow_html=True)
        goal_type="clf" if "Category" in q2 else "reg"

    # Q3 — Class balance (classification only)
    is_imbalanced="no"
    if goal_type=="clf":
        # Auto-hint from last column
        last_col=df.columns[-1]
        if df[last_col].nunique()<=10:
            vc=df[last_col].value_counts(normalize=True)
            auto_hint=f" (Cột cuối '{last_col}': {', '.join(f'{v:.0%} {k}' for k,v in list(vc.items())[:3])})"
        else: auto_hint=""
        st.markdown(f'<div class="quiz-card"><div class="quiz-q">3️⃣  Phân phối các class trong dữ liệu như thế nào?{auto_hint}</div>'
                    '<div class="quiz-hint">Ví dụ: 95% "No Churn" và 5% "Churn" → mất cân bằng nghiêm trọng</div>',
                    unsafe_allow_html=True)
        q3=st.radio("q3",["⚖️  Tương đối cân bằng (không class nào chiếm >80%)",
                           "⚠️  Mất cân bằng — một class chiếm đa số áp đảo",
                           "❓  Chưa biết"],
                    key="adv_q3",label_visibility="collapsed")
        st.markdown('</div>',unsafe_allow_html=True)
        is_imbalanced="yes" if "Mất cân bằng" in q3 else "no"

    # Q4 — Interpretability
    st.markdown('<div class="quiz-card"><div class="quiz-q">4️⃣  Mức độ cần giải thích kết quả (interpretability)?</div>'
                '<div class="quiz-hint">Management hoặc audit có cần hiểu tại sao model đưa ra quyết định đó không?</div>',
                unsafe_allow_html=True)
    q4=st.radio("q4",["🔍  Cao — cần giải thích rõ ràng cho management / audit / compliance",
                       "⚡  Thấp — chỉ cần độ chính xác cao nhất (black-box OK)"],
                key="adv_q4",label_visibility="collapsed")
    st.markdown('</div>',unsafe_allow_html=True)
    interp="high" if "Cao" in q4 else "low"

    # --- Generate ---
    cat_key,recs,size,n=get_recommendations(df,goal_type,interp,is_imbalanced)
    cat_info=ALGO_CATEGORIES[cat_key]

    st.markdown("---")

    # Decision path breadcrumb
    path_nodes=[]
    if goal_type in ("clf","reg"):
        path_nodes.append(('<span class="dp-node dp-blue">Supervised</span>',''))
        if goal_type=="clf":
            path_nodes.append(('<span class="dp-arrow">→</span><span class="dp-node dp-blue">Classification</span>',''))
        else:
            path_nodes.append(('<span class="dp-arrow">→</span><span class="dp-node dp-purple">Regression</span>',''))
    elif goal_type=="clus":
        path_nodes.append(('<span class="dp-node dp-green">Unsupervised</span>',''))
        path_nodes.append(('<span class="dp-arrow">→</span><span class="dp-node dp-green">Clustering</span>',''))
    else:
        path_nodes.append(('<span class="dp-node dp-green">Unsupervised</span>',''))
        path_nodes.append(('<span class="dp-arrow">→</span><span class="dp-node dp-purple">Association</span>',''))
    size_map={"small":f"Small &lt;500 ({n_rows:,} rows)","medium":f"Medium ({n_rows:,} rows)","large":f"Large ({n_rows:,} rows)"}
    path_nodes.append(f'<span class="dp-arrow">→</span><span class="dp-node dp-blue">{size_map.get(size,"")} </span>')
    path_nodes.append(f'<span class="dp-arrow">→</span><span class="dp-node dp-{"amber" if interp=="high" else "green"}">{"High interpretability" if interp=="high" else "High accuracy"}</span>')
    if is_imbalanced=="yes":
        path_nodes.append('<span class="dp-arrow">→</span><span class="dp-node dp-amber">⚠️ Imbalanced classes</span>')

    st.markdown(
        '<div style="font-size:.68rem;font-weight:700;color:#4b5563;text-transform:uppercase;letter-spacing:.07em;margin-bottom:4px">Decision Path</div>'
        '<div class="decision-path">'+"".join(p if isinstance(p,str) else p[0] for p in path_nodes)+'</div>',
        unsafe_allow_html=True)

    # Category highlight card
    st.markdown(
        f'<div style="background:{cat_info["bg"]};border:1px solid {cat_info["border"]};border-radius:11px;padding:13px 16px;margin:10px 0;">'
        f'<div style="display:flex;align-items:center;gap:9px;margin-bottom:5px">'
        f'<span style="font-size:1.15rem">{cat_info["icon"]}</span>'
        f'<span style="font-size:.9rem;font-weight:700;color:{cat_info["color"]}">{cat_info["title"]}</span>'
        f'</div>'
        f'<div style="font-size:.77rem;color:#9ca3af">{cat_info["desc"]}</div>'
        f'<div style="font-size:.72rem;color:#6b7280;margin-top:4px">📌 {cat_info["when"]}</div>'
        f'</div>', unsafe_allow_html=True)

    # Algorithm recommendations
    tier_cfg={
        "prereq":("rb-pre","⚠️ Prerequisite","ri-prereq"),
        "primary":("rb-pri","⭐ Primary","ri-primary"),
        "alt":("rb-alt","🔵 Alternative","ri-alt"),
        "opt":("rb-opt","🟢 Optional","ri-opt"),
    }
    st.markdown('<div style="font-size:.68rem;font-weight:700;color:#4b5563;text-transform:uppercase;letter-spacing:.07em;margin:10px 0 6px">Recommended Algorithms</div>',unsafe_allow_html=True)

    rec_methods=[]
    for mname,tier,why in recs:
        bc,label,rc=tier_cfg.get(tier,tier_cfg["opt"])
        st.markdown(
            f'<div class="rec-item {rc}">'
            f'<span class="rec-badge {bc}">{label}</span>'
            f'<div><div class="rec-name">{mname}</div><div class="rec-why">{why}</div></div>'
            f'</div>', unsafe_allow_html=True)
        if mname in METHODS: rec_methods.append(mname)

    # Apply button
    st.markdown("<br>",unsafe_allow_html=True)
    ca,cb,_=st.columns([1,1,3])
    with ca:
        if st.button("✅ Apply Recommended",type="primary",key="apply_rec"):
            valid=[m for m in rec_methods if m in METHODS]
            st.session_state["selected_methods"]=valid
            st.session_state["rec_applied"]=True
            st.rerun()
    with cb:
        if st.button("🗑️ Clear",key="clr_rec"):
            st.session_state["selected_methods"]=[]; st.session_state["rec_applied"]=False; st.rerun()

    if st.session_state.get("rec_applied") and st.session_state.get("selected_methods"):
        sel=st.session_state["selected_methods"]
        chips=" ".join(f'<span style="background:rgba(59,130,246,.15);color:#60a5fa;padding:2px 9px;border-radius:4px;font-size:.73rem">{m}</span>' for m in sel)
        st.markdown(f'<div style="margin-top:6px;padding:8px 12px;background:#0d1829;border:1px solid #1e3a5f;border-radius:8px;font-size:.78rem;color:#6b7280">✅ Applied: {chips}<br><span style="color:#4b5563;font-size:.71rem">Scroll to "Run ML Models" tab → configure columns → click Run</span></div>',unsafe_allow_html=True)


def render_fitness_report(results, df, goal=""):
    sec_hdr("🏅","Model Fitness Report","amber",subtitle="Grade A–D  ·  Verdict  ·  Actionable Advice")

    def _grade_clf(f1):
        try: f=float(f1)
        except: return "D",0
        if f>=.90: return "A",f
        if f>=.75: return "B",f
        if f>=.60: return "C",f
        return "D",f

    def _grade_reg(r2):
        try: r=float(r2)
        except: return "D",0
        if r>=.80: return "A",r
        if r>=.60: return "B",r
        if r>=.40: return "C",r
        return "D",r

    def _grade_clus(sil):
        try: s=float(sil)
        except: return "D",0
        if s>=.70: return "A",s
        if s>=.50: return "B",s
        if s>=.25: return "C",s
        return "D",s

    verdict_map={
        "A":("✅ Highly suitable — ready for deployment consideration","fv-A"),
        "B":("✅ Suitable — good performance, minor improvements possible","fv-B"),
        "C":("⚠️ Moderate — use with caution; try feature engineering or more data","fv-C"),
        "D":("❌ Not recommended — poor fit; review data quality & feature selection","fv-D"),
    }
    advice_clf={
        "A":"F1 và Accuracy cao. Kiểm tra overfitting bằng cross-validation. Validate trên holdout set trước khi deploy.",
        "B":"Kết quả tốt nhưng còn dư địa. Thử: thêm features, tune hyperparameters, hoặc SMOTE nếu class imbalanced.",
        "C":"Hiệu suất trung bình. Kiểm tra: (1) data quality, (2) class imbalance, (3) feature relevance. Thử Random Forest hoặc thêm features.",
        "D":"Model kém. Nguyên nhân thường: data thiếu chất lượng, features không relevant, hoặc sai loại model. Review lại Step 2 Prep Data.",
    }
    advice_reg={
        "A":"R² tốt — model giải thích được phần lớn variance. Kiểm tra residuals có pattern không. Validate trên out-of-time data.",
        "B":"Fit tốt. Thử thêm interaction terms, polynomial features, hoặc loại bỏ outliers để cải thiện thêm.",
        "C":"Fit yếu. Có thể relationship phi tuyến → thử Neural Networks MLP. Hoặc cần feature engineering.",
        "D":"R² rất thấp — model không giải thích được biến động. Xem xét lại target variable và feature selection.",
    }
    advice_clus={
        "A":"Clusters rõ ràng và tách biệt. Validate với domain expert. Đặt tên cluster theo business meaning.",
        "B":"Separation tốt. Thử K khác nhau hoặc feature selection để cải thiện Silhouette score.",
        "C":"Clusters chồng lấn. Thử: chuẩn hóa features, loại outliers, hoặc thay đổi số K.",
        "D":"Không tìm được clusters rõ ràng. Data có thể không có cấu trúc nhóm tự nhiên — cân nhắc Association Rules.",
    }

    cards=[]
    for r in results:
        method=r.get("Method","Unknown")
        if "F1-Score" in r:
            grade,score=_grade_clf(r.get("F1-Score"))
            score_lbl=f"F1={r.get('F1-Score','?')}  Acc={r.get('Accuracy','?')}  AUC={r.get('AUC','?')}"
            advice=advice_clf[grade]
        elif "R²" in r:
            grade,score=_grade_reg(r.get("R²"))
            score_lbl=f"R²={r.get('R²','?')}  RMSE={r.get('RMSE','?')}"
            advice=advice_reg[grade]
        elif "Silhouette" in r:
            grade,score=_grade_clus(r.get("Silhouette","0"))
            score_lbl=f"Silhouette={r.get('Silhouette','?')}  K={r.get('K','?')}"
            advice=advice_clus[grade]
        else:
            grade,score,score_lbl,advice="C",0.5,"See run output above","Review metrics above."
        verdict_txt,verdict_cls=verdict_map[grade]
        bar_pct=int(min(max(score,0),1)*100)
        cards.append({"method":method,"grade":grade,"score_lbl":score_lbl,
                      "verdict_txt":verdict_txt,"verdict_cls":verdict_cls,"advice":advice,"bar_pct":bar_pct})

    n_c=min(len(cards),3); cols=st.columns(n_c) if n_c>1 else [st.container()]
    for i,card in enumerate(cards):
        with cols[i%n_c] if n_c>1 else cols[0]:
            st.markdown(
                f'<div class="fitness-card">'
                f'<div style="display:flex;align-items:baseline;gap:10px;margin-bottom:3px">'
                f'<div class="fitness-grade fg-{card["grade"]}">{card["grade"]}</div>'
                f'<div><div class="fitness-method">{card["method"]}</div>'
                f'<div class="fitness-score">{card["score_lbl"]}</div></div></div>'
                f'<div class="fitness-bar-wrap"><div class="fitness-bar fb-{card["grade"]}" style="width:{card["bar_pct"]}%"></div></div>'
                f'<div class="fitness-verdict {card["verdict_cls"]}">{card["verdict_txt"]}</div>'
                f'<div class="fitness-advice">{card["advice"]}</div>'
                f'</div>', unsafe_allow_html=True)

    if len(cards)>1:
        grade_order={"A":4,"B":3,"C":2,"D":1}
        best=max(cards,key=lambda x:grade_order.get(x["grade"],0))
        st.markdown(
            f'<div style="background:rgba(59,130,246,.07);border:1px solid rgba(59,130,246,.22);'
            f'border-radius:10px;padding:11px 16px;margin-top:10px;font-size:.81rem;">'
            f'🏆 <b style="color:#60a5fa">Best model: {best["method"]}</b> (Grade {best["grade"]}) — '
            f'<span style="color:#9ca3af">{best["verdict_txt"]}</span>'
            f'</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULTS={
    "sheets":{},"active_sheet":None,"wizard_step":1,"ai_blueprint":"",
    "prep_transforms":{},"prep_log":{},"enc_mapping":{},"user_goal":"",
    "selected_methods":[],"run_results":[],"run_exports":{},"comparison_ai":"",
    "gemini_key":"","openrouter_key":"","ai_vn":"","rec_applied":False,
    "chat_messages":[], "_trained_models":{}, "_split_data":{},"_run_figures":{},"_file_cache":{},
}
for k,v in DEFAULTS.items():
    if k not in st.session_state: st.session_state[k]=v

# ═══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def render_stepper(ws):
    steps=["Load Data","Prep Data","Analysis & Profiling","Advise & Predict"]
    parts=[]
    for i,s in enumerate(steps):
        n=i+1
        if n<ws:    cc,lc,sym="sc-d","sl-d","✓"
        elif n==ws: cc,lc,sym="sc-a","sl-a",str(n)
        else:       cc,lc,sym="sc-p","sl-p",str(n)
        parts.append(f'<div class="step-item"><div class="step-circle {cc}">{sym}</div><div class="step-lbl {lc}">{s}</div></div>')
        if i<len(steps)-1:
            conn_cls="sc-conn-d" if n<ws else "sc-conn-p"
            parts.append(f'<div class="step-conn {conn_cls}"></div>')
    st.markdown('<div class="stepper">'+"".join(parts)+'</div>',unsafe_allow_html=True)

def sec_hdr(icon,title,color="blue",subtitle=""):
    sub="<span class='sec-sub'>"+subtitle+"</span>" if subtitle else ""
    st.markdown(
        f'<div class="sec-hdr"><div class="sec-icon si-{color}">{icon}</div>'
        f'<span class="sec-title">{title}</span>{sub}</div>',
        unsafe_allow_html=True)

def stat_row(items):
    cols=st.columns(len(items))
    for col,(val,lbl,cls) in zip(cols,items):
        with col:
            st.markdown(f'<div class="stat-card"><div class="stat-val {cls}">{val}</div><div class="stat-lbl">{lbl}</div></div>',unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style="padding:12px 0 14px">
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:4px">
        <div style="width:36px;height:36px;background:linear-gradient(135deg,#1d4ed8,#7c3aed);border-radius:9px;display:flex;align-items:center;justify-content:center;font-size:1.2rem">🧬</div>
        <div><div style="font-size:1rem;font-weight:700;color:#f9fafb">DataMine AI</div>
        <div style="font-size:.68rem;color:#4b5563">Analytics Platform</div></div>
      </div>
    </div>""",unsafe_allow_html=True)

    ws_sb=st.session_state["wizard_step"]
    steps_sb=["Load Data","Prep Data","Analysis & Profiling","Advise & Predict"]
    st.markdown('<div style="font-size:.63rem;font-weight:700;letter-spacing:.1em;color:#374151;text-transform:uppercase;margin-bottom:5px">WORKFLOW</div>',unsafe_allow_html=True)
    for i,s in enumerate(steps_sb):
        n=i+1
        if n<ws_sb:   nc,tc,ac="sn-done","sbt-d",""; sym="✓"
        elif n==ws_sb: nc,tc,ac="sn-active","sbt-a"," sb-step-a"; sym=str(n)
        else:          nc,tc,ac="sn-pend","sbt-p",""; sym=str(n)
        st.markdown(f'<div class="sb-step{ac}"><div class="sb-num {nc}">{sym}</div><div class="sb-txt {tc}">{s}</div></div>',unsafe_allow_html=True)

    st.markdown('<div style="border-top:1px solid #1f2937;padding-top:10px;margin-top:12px;font-size:.63rem;font-weight:700;letter-spacing:.1em;color:#374151;text-transform:uppercase;margin-bottom:7px">DATA UPLOAD</div>',unsafe_allow_html=True)

    # Header row selector
    _header_row = st.number_input(
        "Dòng header (0 = dòng đầu tiên)",
        min_value=0, max_value=20, value=st.session_state.get("_header_row", 0), step=1,
        key="_header_row_input",
        help="Tăng số này nếu file có logo/tiêu đề phụ ở đầu. Nhấn 'Cập nhật header' sau khi thay đổi."
    )
    st.session_state["_header_row"] = int(_header_row)

    uploaded_files = st.file_uploader("Drop files here", type=["csv","xlsx","xls","xlsm","json","txt"],
        accept_multiple_files=True, label_visibility="collapsed")
    if uploaded_files:
        for uf in uploaded_files:
            # Nếu file đã được cache (đã load trước đó), KHÔNG load lại
            # Tránh bị override sheet khi user đang chọn sheet khác
            already_cached = uf.name in st.session_state.get("_file_cache", {})
            already_in_sheets = any(
                k == uf.name or k.startswith(uf.name + "::")
                for k in st.session_state.get("sheets", {})
            )
            if already_cached and already_in_sheets:
                continue  # Skip — sheet selector sẽ handle việc đổi sheet

            new = load_file(uf, header_row=int(_header_row))
            for k, v in new.items():
                if k not in st.session_state["sheets"]:
                    st.session_state["sheets"][k] = v
                    if st.session_state["active_sheet"] is None:
                        st.session_state["active_sheet"] = k

    all_sheets = st.session_state.get("sheets", {})
    if all_sheets:
        keys = list(all_sheets.keys())

        # ── Active file selector (chỉ hiện khi nhiều file) ───────────────
        active_key = st.session_state["active_sheet"] or keys[0]
        if active_key not in keys:
            active_key = keys[0]
            st.session_state["active_sheet"] = active_key

        if len(keys) > 1:
            # Hiển thị tên đẹp hơn (bỏ phần ::SheetName nếu có)
            def _display_key(k):
                parts = k.split("::")
                return f"📄 {parts[0]}" if len(parts)==1 else f"📄 {parts[0]}  ›  {parts[1]}"
            disp_keys = [_display_key(k) for k in keys]
            cur_idx = keys.index(active_key) if active_key in keys else 0
            chosen_disp = st.selectbox("File đang phân tích", disp_keys,
                                       index=cur_idx, label_visibility="visible")
            chosen = keys[disp_keys.index(chosen_disp)]
            if chosen != active_key:
                st.session_state["active_sheet"] = chosen
                active_key = chosen
                st.rerun()

        # ── Sheet selector — luôn hiển thị nếu Excel có nhiều sheet ──────
        file_name = active_key.split("::")[0] if "::" in active_key else active_key
        cache = st.session_state.get("_file_cache", {}).get(file_name)

        if cache and len(cache.get("sheet_names", [])) > 1:
            cur_sheet = cache.get("current_sheet", cache["sheet_names"][0])
            sheet_names = cache["sheet_names"]
            sheet_idx = sheet_names.index(cur_sheet) if cur_sheet in sheet_names else 0

            st.markdown('<div style="font-size:.7rem;color:#6b7280;margin-top:8px;margin-bottom:3px">📋 CHỌN SHEET</div>',
                        unsafe_allow_html=True)
            sel_sheet = st.selectbox(
                "Sheet trong file", sheet_names, index=sheet_idx,
                label_visibility="collapsed",
                help=f"File có {len(sheet_names)} sheets — chọn sheet để đọc dữ liệu"
            )

            # Hiển thị danh sách sheet nhỏ
            sheet_chips = " ".join(
                f'<span style="background:{"rgba(59,130,246,.2)" if s==sel_sheet else "var(--card2)"};'
                f'color:{"#60a5fa" if s==sel_sheet else "#6b7280"};padding:2px 8px;border-radius:4px;'
                f'font-size:.7rem;margin:1px">{s}</span>'
                for s in sheet_names
            )
            st.markdown(f'<div style="margin-bottom:4px">{sheet_chips}</div>', unsafe_allow_html=True)

            if sel_sheet != cur_sheet:
                try:
                    with st.spinner(f"Đang đọc sheet '{sel_sheet}'..."):
                        new_df = load_sheet(cache["raw_bytes"], sel_sheet,
                                            cache["engine"], int(_header_row))
                    new_key = f"{file_name}::{sel_sheet}"
                    # Xóa key cũ, thêm key mới
                    st.session_state["sheets"].pop(active_key, None)
                    st.session_state["sheets"][new_key] = new_df
                    st.session_state["active_sheet"] = new_key
                    cache["current_sheet"] = sel_sheet
                    # Reset prep transforms vì đổi sheet
                    st.session_state["prep_transforms"].pop(active_key, None)
                    st.session_state["prep_log"].pop(active_key, None)
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Lỗi đọc sheet '{sel_sheet}': {e}")
        elif cache and len(cache.get("sheet_names", [])) == 1:
            st.markdown(
                f'<div style="font-size:.72rem;color:#4b5563;padding:3px 0">'
                f'📋 Sheet: <b style="color:#9ca3af">{cache["sheet_names"][0]}</b></div>',
                unsafe_allow_html=True
            )

        if len(all_sheets) > 1:
            with st.expander("🔗 Merge Files", expanded=False):
                left_k = st.selectbox("Left file", keys, key="ml")
                right_k = st.selectbox("Right file", [k for k in keys if k != left_k], key="mr")
                shared = list(set(all_sheets[left_k].columns) & set(all_sheets[right_k].columns)) if right_k else []
                join_on = st.selectbox("Join on", shared or ["(no shared columns)"], key="mo")
                join_type = st.selectbox("Join type", ["inner","left","outer","right"], key="mt")
                if shared and st.button("Merge →", key="dm"):
                    merged = pd.merge(all_sheets[left_k], all_sheets[right_k], on=join_on, how=join_type)
                    mk = f"merged_{left_k[:12]}+{right_k[:12]}"
                    st.session_state["sheets"][mk] = merged
                    st.session_state["active_sheet"] = mk
                    st.success(f"Merged → {mk} ({merged.shape[0]:,} rows)")
                    st.rerun()

    st.markdown('<div style="border-top:1px solid #1f2937;padding-top:10px;margin-top:12px;font-size:.63rem;font-weight:700;letter-spacing:.1em;color:#374151;text-transform:uppercase;margin-bottom:7px">AI KEYS</div>',unsafe_allow_html=True)
    st.text_input("Gemini key",type="password",key="gemini_key",placeholder="AIza…",label_visibility="collapsed")
    st.caption("Gemini (aistudio.google.com)")
    st.text_input("OpenRouter key",type="password",key="openrouter_key",placeholder="sk-or-…",label_visibility="collapsed")
    st.caption("OpenRouter (openrouter.ai) — fallback")
    g_ok=bool(st.session_state.get("gemini_key","").strip())
    o_ok=bool(st.session_state.get("openrouter_key","").strip())
    if g_ok or o_ok:
        st.success(f"AI ready: {'Gemini' if g_ok else 'OpenRouter'}")
        if st.button("🔑 Test kết nối AI", key="test_ai_key"):
            with st.spinner("Đang kiểm tra..."):
                test_result = ask_ai("Trả lời đúng 5 chữ: 'Kết nối thành công!'")
            if "❌" in test_result or "⚠️" in test_result:
                st.error(test_result)
            else:
                st.success(f"✅ {test_result[:80]}")
    else:
        st.markdown('<div style="font-size:.74rem;color:#4b5563;padding:4px 0">⚪ No key — ML runs fine; AI explanations disabled.</div>',unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
ws=st.session_state["wizard_step"]

st.markdown(
    '<div class="app-header">'
    '<div class="app-logo">🧬</div>'
    '<div><div class="app-title">DataMine AI</div>'
    '<div class="app-sub">Upload · Prep · Profile · Predict · Chat</div></div>'
    '<div class="version-badge">v2.0</div></div>',
    unsafe_allow_html=True)

# ── TOP-LEVEL NAVIGATION ────────────────────────────────────────────────────
_app_nav=st.radio("Chế độ",["📊 Workflow","💬 AI Chat"],horizontal=True,
                  label_visibility="collapsed",key="app_nav")

if _app_nav=="💬 AI Chat":
    g_ok_nav=bool(st.session_state.get("gemini_key","").strip())
    o_ok_nav=bool(st.session_state.get("openrouter_key","").strip())

    if not g_ok_nav and not o_ok_nav:
        st.warning("⚠️ Vui lòng nhập **Gemini API Key** trong sidebar để bắt đầu chat.")
        st.stop()

    # ── Context badge ────────────────────────────────────────────────────────
    _has_data = bool(st.session_state.get("sheets", {}))
    _has_res  = bool(st.session_state.get("run_results", []))
    _ctx_parts = []
    if _has_data: _ctx_parts.append(f"📊 {len(st.session_state['sheets'])} file(s) đã upload")
    if _has_res:  _ctx_parts.append(f"🤖 {len(st.session_state['run_results'])} kết quả model")
    if _ctx_parts:
        st.markdown(
            f'<div style="background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.2);'
            f'border-radius:8px;padding:8px 14px;font-size:.78rem;color:#6ee7b7;margin-bottom:12px">'
            f'✅ AI đang biết: {" · ".join(_ctx_parts)}</div>', unsafe_allow_html=True)
    else:
        st.markdown(
            '<div style="background:rgba(99,102,241,.08);border:1px solid rgba(99,102,241,.2);'
            'border-radius:8px;padding:8px 14px;font-size:.78rem;color:#a5b4fc;margin-bottom:12px">'
            '💡 Chưa có data — có thể hỏi kiến thức data science tổng quát, '
            'hoặc upload data ở tab Workflow rồi quay lại đây.</div>', unsafe_allow_html=True)

    # ── Suggestion chips ─────────────────────────────────────────────────────
    _SUGGESTIONS = [
        ("🤖", "Giải thích sự khác nhau giữa Random Forest và Logistic Regression"),
        ("📊", "Làm thế nào để xử lý dữ liệu bị mất (missing values)?"),
        ("🎯", "Khi nào nên dùng K-Means, khi nào dùng Hierarchical Clustering?"),
        ("⚠️", "Class imbalance là gì và cách xử lý?"),
        ("📈", "Tóm tắt và phân tích dữ liệu tôi đã upload"),
        ("🏆", "Model nào trong kết quả vừa chạy là tốt nhất và tại sao?"),
    ]
    sec_hdr("💬", "AI Chat", "purple", subtitle="Hỏi bất kỳ điều gì về dữ liệu hoặc data science")
    st.markdown('<div style="font-size:.75rem;color:#6b7280;margin-bottom:8px">💡 Gợi ý:</div>', unsafe_allow_html=True)
    _sc = st.columns(3)
    for _i, (_icon, _txt) in enumerate(_SUGGESTIONS):
        with _sc[_i % 3]:
            if st.button(f"{_icon} {_txt[:40]}…" if len(_txt) > 40 else f"{_icon} {_txt}",
                         key=f"sug_{_i}", use_container_width=True):
                st.session_state["_chat_pending"] = _txt
                st.rerun()

    st.markdown("---")

    # ── Input box ───────────────────────────────────────────────────────────
    with st.form("chat_form", clear_on_submit=True):
        _fc, _bc = st.columns([5, 1])
        with _fc:
            _user_q = st.text_input("Câu hỏi", placeholder="Nhập câu hỏi rồi nhấn Gửi…",
                                    label_visibility="collapsed")
        with _bc:
            _submitted = st.form_submit_button("📨 Gửi", use_container_width=True)

    # Xử lý câu hỏi từ suggestion hoặc form
    _pending = st.session_state.pop("_chat_pending", None)
    _question = _pending or (_user_q.strip() if _submitted and _user_q.strip() else None)

    if _question:
        st.session_state["chat_messages"].append({"role": "user", "content": _question})
        _data_ctx = _build_data_context()
        _sys_prompt = (
            "Bạn là chuyên gia data science và data mining. "
            "Trả lời bằng tiếng Việt, rõ ràng, có ví dụ cụ thể. "
            "Nếu có dữ liệu, hãy tham chiếu cột/số liệu thực tế.\n\n"
            + (f"DỮ LIỆU NGƯỜI DÙNG:\n{_data_ctx[:4000]}" if _has_data
               else "Chưa có dữ liệu — trả lời kiến thức tổng quát.")
        )
        _full_prompt = _sys_prompt + "\n\n---\nCÂU HỎI: " + _question
        with st.spinner("⏳ AI đang suy nghĩ..."):
            _answer = ask_ai(_full_prompt)
        st.session_state["chat_messages"].append({"role": "assistant", "content": _answer})
        st.rerun()

    # ── Lịch sử chat ────────────────────────────────────────────────────────
    _history = st.session_state.get("chat_messages", [])
    if not _history:
        st.markdown(
            '<div style="text-align:center;padding:40px 20px">'
            '<div style="font-size:2.5rem">🤖</div>'
            '<div style="color:#4b5563;font-size:.85rem;margin-top:8px">'
            'Chưa có câu hỏi nào. Chọn gợi ý hoặc nhập câu hỏi ở trên.</div></div>',
            unsafe_allow_html=True)
    else:
        for _msg in reversed(_history):  # mới nhất ở trên
            if _msg["role"] == "user":
                st.markdown(
                    f'<div style="background:rgba(59,130,246,.12);border:1px solid rgba(59,130,246,.25);'
                    f'border-radius:10px;padding:10px 14px;margin:6px 0;font-size:.88rem">'
                    f'🧑 <b>Bạn:</b> {_msg["content"]}</div>', unsafe_allow_html=True)
            else:
                render_ai_box(_msg["content"], "🤖 AI")
        if st.button("🗑️ Xoá lịch sử", key="clr_chat"):
            st.session_state["chat_messages"] = []
            st.rerun()

    st.stop()

# ── WORKFLOW MODE ────────────────────────────────────────────────────────────
render_stepper(ws)

if not st.session_state["sheets"]:
    st.markdown("""
    <div class="welcome">
      <h2>Welcome to DataMine AI</h2>
      <p>A guided data analytics platform. Upload your data and follow 4 steps to clean, profile, model, and get AI-driven predictions.</p>
      <div class="step-list">
        <div class="step-list-item"><span class="sli-num">01</span><div class="sli-text"><b>Load Data</b>Upload CSV, Excel, JSON or TXT. Multiple files supported — merge later.</div></div>
        <div class="step-list-item"><span class="sli-num">02</span><div class="sli-text"><b>Prep Data</b>Auto-detect and fix missing values, duplicates, outliers and encoding.</div></div>
        <div class="step-list-item"><span class="sli-num">03</span><div class="sli-text"><b>Analysis & Profiling</b>Explore distributions, correlations, missing value charts. Run ML models side-by-side.</div></div>
        <div class="step-list-item"><span class="sli-num">04</span><div class="sli-text"><b>Advise & Predict</b>AI blueprint + model comparison + actionable recommendations. Export to Excel.</div></div>
      </div>
    </div>""",unsafe_allow_html=True)
    st.markdown('<div style="text-align:center;color:#4b5563;font-size:.83rem;padding:6px">👈 Upload a file in the sidebar to begin</div>',unsafe_allow_html=True)
    st.stop()

active_name=st.session_state["active_sheet"]
raw_df=st.session_state["sheets"][active_name]
_prep_stack=st.session_state["prep_transforms"].get(active_name,[])
df_active=_json_to_df(_prep_stack[-1][1]) if _prep_stack else raw_df.copy()

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════════
sec_hdr("📂","Step 1 — Load Data","blue",subtitle=f"{len(st.session_state['sheets'])} file(s) loaded")

all_loaded=st.session_state.get("sheets",{})
miss_pct=f"{raw_df.isnull().mean().mean():.1%}"; dup_count=raw_df.duplicated().sum()

stat_row([
    (f"{len(all_loaded)}","Files Loaded","sv-blue"),
    (f"{sum(d.shape[0] for d in all_loaded.values()):,}","Total Rows","sv-gray"),
    (miss_pct,"Missing Rate","sv-amber" if raw_df.isnull().mean().mean()>.1 else "sv-green"),
    (f"{dup_count:,}","Duplicates","sv-red" if dup_count>0 else "sv-green"),
])

# ── Nút cập nhật header — đặt ngay trên preview data ────────────────────────
_cur_header_s1 = st.session_state.get("_header_row", 0)
_h_col, _h_btn = st.columns([4, 1])
with _h_col:
    st.markdown(
        f'<div style="font-size:.8rem;color:#6b7280;padding:4px 0">'
        f'📌 Dòng header: <b style="color:#f9fafb">{_cur_header_s1}</b> '
        f'(đổi số ở sidebar → nhấn "Cập nhật header" để reload)</div>',
        unsafe_allow_html=True)
with _h_btn:
    if st.button("🔄 Cập nhật header", key="update_header_s1"):
        _active_k = st.session_state.get("active_sheet","")
        _fname = _active_k.split("::")[0] if "::" in _active_k else _active_k
        _cache_s1 = st.session_state.get("_file_cache", {}).get(_fname)
        if _cache_s1:
            _cur_sh_s1 = _cache_s1.get("current_sheet", _cache_s1["sheet_names"][0])
            try:
                _new_df_s1 = load_sheet(_cache_s1["raw_bytes"], _cur_sh_s1,
                                        _cache_s1["engine"], _cur_header_s1)
                st.session_state["sheets"][_active_k] = _new_df_s1
                st.session_state["prep_transforms"].pop(_active_k, None)
                st.session_state["prep_log"].pop(_active_k, None)
                st.success(f"✅ Đã đọc lại từ dòng header {_cur_header_s1}")
                st.rerun()
            except Exception as _he_s1:
                st.error(f"Lỗi: {_he_s1}")
        else:
            st.info("Chỉ hỗ trợ file Excel. Với CSV, upload lại với header mới.")

file_tabs=st.tabs([f"📄 {n}" for n in all_loaded.keys()])
for tab,(sname,sdf) in zip(file_tabs,all_loaded.items()):
    _s=st.session_state["prep_transforms"].get(sname,[])
    disp=_json_to_df(_s[-1][1]) if _s else sdf
    with tab:
        st.caption(f"{disp.shape[0]:,} rows × {disp.shape[1]} cols"+(" · ✅ Cleaned" if _s else " · Original"))
        t1,t2,t3=st.tabs(["Table","Statistics","Column Types"])
        with t1: st.dataframe(_sanitize_df(disp.head(50)),width='stretch')
        with t2: st.dataframe(_sanitize_df(disp.describe(include="all")),width='stretch')
        with t3:
            dt=disp.dtypes.reset_index(); dt.columns=["Column","Type"]
            dt["Nulls"]=disp.isnull().sum().values
            dt["Null%"]=(disp.isnull().mean()*100).round(1).values
            dt["Unique"]=disp.nunique().values
            dt["Sample"]=[str(disp[c].dropna().iloc[0]) if disp[c].dropna().shape[0]>0 else "" for c in disp.columns]
            st.dataframe(_sanitize_df(dt),width='stretch')

if ws==1:
    st.markdown("<br>",unsafe_allow_html=True)
    c1,_=st.columns([1,4])
    with c1:
        if st.button("Continue to Prep Data →",type="primary"):
            st.session_state["wizard_step"]=2; st.rerun()
st.divider()

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 — PREP DATA
# ═══════════════════════════════════════════════════════════════════════════════
if ws>=2:
    sec_hdr("🧹","Step 2 — Prep Data","green",subtitle="Auto-detect & fix data quality issues")
    all_loaded_2=st.session_state.get("sheets",{})
    if len(all_loaded_2)>1:
        st.markdown('<div style="background:rgba(59,130,246,.08);border:1px solid rgba(59,130,246,.2);border-radius:8px;padding:8px 14px;font-size:.8rem;color:#93c5fd;margin-bottom:10px">💡 Multiple files — select which one to clean.</div>',unsafe_allow_html=True)
        prep_target=st.selectbox("Prepare which file?",list(all_loaded_2.keys()),
            index=list(all_loaded_2.keys()).index(active_name) if active_name in all_loaded_2 else 0,
            key="prep_sel")
    else:
        prep_target=active_name
    prep_raw=all_loaded_2.get(prep_target,raw_df)

    if prep_target not in st.session_state["prep_log"]: st.session_state["prep_log"][prep_target]=[]
    if prep_target not in st.session_state["prep_transforms"]: st.session_state["prep_transforms"][prep_target]=[]
    stack=st.session_state["prep_transforms"][prep_target]; log=st.session_state["prep_log"][prep_target]
    cur_json=stack[-1][1] if stack else _df_to_json(prep_raw); cur_df=_json_to_df(cur_json)

    st.markdown(
        f'<div style="background:#0f1629;border:1px solid #1f2937;border-radius:8px;padding:7px 14px;font-size:.8rem;color:#6b7280;margin-bottom:12px">'
        f'Working on: <b style="color:#60a5fa">{prep_target}</b> — {cur_df.shape[0]:,} rows × {cur_df.shape[1]} cols'
        f'{"  <span style=color:#34d399>✅ Cleaned</span>" if stack else "  · Original"}</div>',
        unsafe_allow_html=True)

    sec_hdr("🔍","Auto-Detected Issues","amber")
    with st.spinner("Scanning data quality…"):
        problems=detect_problems(cur_json)
    sev_map={"err":("pc-err","🔴"),"warn":("pc-warn","🟡"),"info":("pc-info","🔵"),"ok":("pc-ok","✅")}
    for p in problems:
        cls,ico=sev_map.get(p["sev"],("pc-info","ℹ️"))
        fix_html=f'<div class="prob-fix">→ {p["fix"]}</div>' if p["fix"] else ""
        st.markdown(
            f'<div class="prob-row">{ico} <span class="prob-chip {cls}">{p["type"].upper()}</span>'
            f'<div><div class="prob-txt">{p["msg"]}</div>{fix_html}</div></div>',
            unsafe_allow_html=True)

    st.markdown('<div style="margin:12px 0 7px;font-size:.79rem;font-weight:600;color:#9ca3af">Apply Fixes →</div>',unsafe_allow_html=True)
    has_missing=cur_df.isnull().any().any(); has_dups=cur_df.duplicated().any()
    has_text=len(cur_df.select_dtypes(include=["object","string"]).columns)>0
    num_c2=cur_df.select_dtypes(include=[np.number]).columns
    has_out=any(((cur_df[c]<cur_df[c].quantile(.25)-3*(cur_df[c].quantile(.75)-cur_df[c].quantile(.25)))|
        (cur_df[c]>cur_df[c].quantile(.75)+3*(cur_df[c].quantile(.75)-cur_df[c].quantile(.25)))).any()
        for c in num_c2 if (cur_df[c].quantile(.75)-cur_df[c].quantile(.25))>0)
    b1,b2,b3,b4=st.columns(4)
    with b1:
        if st.button("🩹 Fix Missing" if has_missing else "✅ No Missing",disabled=not has_missing,key=f"bm_{prep_target[:8]}"):
            nj,msg=fix_missing(cur_json); stack.append(("🩹 Fix Missing",nj)); log.append(("🩹 Missing",msg))
            detect_problems.clear(); st.rerun()
    with b2:
        if st.button("🗑️ Remove Dups" if has_dups else "✅ No Duplicates",disabled=not has_dups,key=f"bd_{prep_target[:8]}"):
            nj,msg=fix_duplicates(cur_json); stack.append(("🗑️ Remove Dups",nj)); log.append(("🗑️ Duplicates",msg))
            detect_problems.clear(); st.rerun()
    with b3:
        if has_out:
            _out_action = st.selectbox("Outlier action",
                ["📐 Cap (giữ dòng, đổi giá trị)","🗑️ Clean (xóa dòng outlier)"],
                key=f"out_act_{prep_target[:8]}", label_visibility="collapsed")
            if st.button("▶ Xử lý Outlier", key=f"bo_{prep_target[:8]}"):
                if "Cap" in _out_action:
                    nj,msg=fix_outliers(cur_json); stack.append(("📐 Cap Outliers",nj)); log.append(("📐 Cap Outliers",msg))
                else:
                    nj,msg=remove_outliers(cur_json); stack.append(("🗑️ Clean Outliers",nj)); log.append(("🗑️ Clean Outliers",msg))
                detect_problems.clear(); st.rerun()
        else:
            st.button("✅ No Extremes", disabled=True, key=f"bo_{prep_target[:8]}")
    with b4:
        if st.button("🔤 Encode Text" if has_text else "✅ No Text Cols",disabled=not has_text,key=f"be_{prep_target[:8]}"):
            nj,msg,mapping=fix_encode(cur_json); stack.append(("🔤 Encode Text",nj)); log.append(("🔤 Encoding",msg))
            st.session_state["enc_mapping"][active_name]=mapping; detect_problems.clear(); st.rerun()

    if st.session_state["enc_mapping"].get(active_name):
        with st.expander("🗂️ Encoding Map",expanded=False):
            for col,vm in st.session_state["enc_mapping"][active_name].items():
                st.markdown(f"**`{col}`**")
                mdf=pd.DataFrame(list(vm.items()),columns=["Original","Encoded"]).sort_values("Encoded")
                st.dataframe(_sanitize_df(mdf),width='stretch',height=min(200,35*len(mdf)+38))
    # ── Grouping & Binning (trong Prep Data) ────────────────────────────────
    st.markdown("---")
    sec_hdr("🗂️","Grouping & Binning","amber",subtitle="Tạo cột mới: nhóm giá trị text hoặc chia khoảng số")
    _gb_c1, _gb_c2 = st.columns([2,3])
    with _gb_c1:
        gb_col=st.selectbox("Chọn cột",["(chọn cột)"] + cur_df.columns.tolist(),key=f"gb_col_{prep_target[:8]}")
    if gb_col and gb_col != "(chọn cột)":
        with _gb_c2:
            gb_type=st.radio("Loại","Binning (số→khoảng) | Grouping (text→nhóm)".split(" | "),
                horizontal=True,key=f"gb_type_{prep_target[:8]}")
        if "Binning" in gb_type and pd.api.types.is_numeric_dtype(cur_df[gb_col]):
            _gc1,_gc2,_gc3=st.columns([1,2,2])
            with _gc1: gb_bins=st.number_input("Số khoảng",2,50,5,key=f"gb_bins_{prep_target[:8]}")
            with _gc2: gb_labels_raw=st.text_input("Nhãn khoảng (cách phẩy, để trống=tự động)",key=f"gb_labels_{prep_target[:8]}")
            with _gc3: gb_new_col=st.text_input("Tên cột mới",value=f"{gb_col}_bin",key=f"gb_new_col_{prep_target[:8]}")
            if st.button("✅ Áp dụng Binning",key=f"do_bin_{prep_target[:8]}"):
                try:
                    labels=None
                    if gb_labels_raw.strip():
                        labels=[l.strip() for l in gb_labels_raw.split(",")]
                        if len(labels)!=gb_bins: labels=None; st.warning(f"Cần {gb_bins} nhãn, bỏ qua labels.")
                    cur_df[gb_new_col]=pd.cut(cur_df[gb_col],bins=int(gb_bins),labels=labels)
                    cur_df[gb_new_col]=cur_df[gb_new_col].astype(str)
                    new_json=_df_to_json(cur_df)
                    st.session_state["prep_transforms"][prep_target].append((f"Binning {gb_col}→{gb_bins} bins",new_json))
                    st.success(f"✅ Đã tạo cột '{gb_new_col}' với {gb_bins} khoảng"); st.rerun()
                except Exception as e: st.error(f"Lỗi Binning: {e}")
        elif "Grouping" in gb_type and not pd.api.types.is_numeric_dtype(cur_df[gb_col]):
            unique_vals=sorted(cur_df[gb_col].dropna().unique().tolist())
            st.markdown(f"**Giá trị duy nhất ({len(unique_vals)}):** {', '.join(str(v) for v in unique_vals[:20])}")
            gb_map_raw=st.text_area("Mapping (mỗi dòng: giá_trị_gốc → nhóm_mới)",
                placeholder="Hà Nội → Miền Bắc\nTP.HCM → Miền Nam",key=f"gb_map_{prep_target[:8]}")
            gb_new_col2=st.text_input("Tên cột mới",value=f"{gb_col}_group",key=f"gb_new_col2_{prep_target[:8]}")
            if st.button("✅ Áp dụng Grouping",key=f"do_group_{prep_target[:8]}"):
                try:
                    mapping={}
                    for line in gb_map_raw.strip().split("\n"):
                        if "→" in line:
                            k,v=line.split("→",1); mapping[k.strip()]=v.strip()
                    cur_df[gb_new_col2]=cur_df[gb_col].map(mapping).fillna(cur_df[gb_col])
                    new_json=_df_to_json(cur_df)
                    st.session_state["prep_transforms"][prep_target].append((f"Grouping {gb_col}→{gb_new_col2}",new_json))
                    st.success(f"✅ Đã tạo cột '{gb_new_col2}' với {cur_df[gb_new_col2].nunique()} nhóm"); st.rerun()
                except Exception as e: st.error(f"Lỗi Grouping: {e}")
        else:
            st.info("Binning chỉ áp dụng cho cột số. Grouping chỉ áp dụng cho cột text.")

    if log:
        st.markdown("---"); sec_hdr("📋","Change Log","green")
        for name2,msg in log:
            st.markdown(f'<div class="log-item"><b>{name2}:</b> {msg}</div>',unsafe_allow_html=True)
        cur_df2=_json_to_df(stack[-1][1]); ca,cb=st.columns(2)
        with ca:
            st.markdown('<div class="diff-label dl-red">📌 Before</div>',unsafe_allow_html=True)
            st.markdown(f'<div class="diff-before"><div style="font-size:.79rem;color:#d1d5db">📏 {prep_raw.shape[0]:,}×{prep_raw.shape[1]} &nbsp;|&nbsp; ❓ {prep_raw.isnull().sum().sum():,} missing &nbsp;|&nbsp; 📋 {prep_raw.duplicated().sum():,} dups</div></div>',unsafe_allow_html=True)
            st.dataframe(_sanitize_df(prep_raw.head(5)),width='stretch',height=165)
        with cb:
            st.markdown('<div class="diff-label dl-green">✅ After</div>',unsafe_allow_html=True)
            st.markdown(f'<div class="diff-after"><div style="font-size:.79rem;color:#d1d5db">📏 {cur_df2.shape[0]:,}×{cur_df2.shape[1]} &nbsp;|&nbsp; ❓ {cur_df2.isnull().sum().sum():,} missing &nbsp;|&nbsp; 📋 {cur_df2.duplicated().sum():,} dups</div></div>',unsafe_allow_html=True)
            st.dataframe(_sanitize_df(cur_df2.head(5)),width='stretch',height=165)
        # ── Xuất Before/After dạng Excel ─────────────────────────────────
        _prep_fname = prep_target.replace('::','_').replace(' ','_')[:30]
        _export_sheets = {
            "Raw Data (Before)": _sanitize_df(prep_raw.copy()),
            "Cleaned Data (After)": _sanitize_df(cur_df2.copy()),
        }
        # Thêm prep log
        if log:
            _log_df = pd.DataFrame([(i+1, step, msg) for i,(step,msg) in enumerate(log)],
                                   columns=["Step","Action","Detail"])
            _export_sheets["Prep Log"] = _log_df
        _before_after_bytes = export_to_excel(_export_sheets)
        st.download_button(
            "⬇️ Download Before / After (Excel)",
            data=_before_after_bytes,
            file_name=f"before_after_{_prep_fname}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        cu,cr=st.columns(2)
        with cu:
            if st.button("↩️ Undo"): stack.pop(); log.pop(); detect_problems.clear(); st.rerun()
        with cr:
            if st.button("🔄 Reset All"):
                st.session_state["prep_transforms"][active_name]=[]; st.session_state["prep_log"][active_name]=[]
                if active_name in st.session_state["enc_mapping"]: del st.session_state["enc_mapping"][active_name]
                detect_problems.clear(); st.rerun()
    else:
        st.info("No fixes applied yet. Click buttons above or proceed if data is already clean.")

    if ws==2:
        st.markdown("<br>",unsafe_allow_html=True)
        c1,_=st.columns([1,4])
        with c1:
            if st.button("Continue to Analysis →",type="primary"):
                st.session_state["wizard_step"]=3; st.rerun()
    st.divider()

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 — ANALYSIS & PROFILING
# ═══════════════════════════════════════════════════════════════════════════════
if ws>=3:
    stack3=st.session_state["prep_transforms"].get(active_name,[])
    df_active=_json_to_df(stack3[-1][1]) if stack3 else raw_df.copy()
    sec_hdr("⚡","Step 3 — Analysis & Profiling","purple",
            subtitle=f"{df_active.shape[0]:,} rows × {df_active.shape[1]} cols")

    adv_tab,eda_tab,run_tab=st.tabs(["🎯 Algorithm Advisor","🔬 EDA & Validation","🤖 Run ML Models"])

    with adv_tab:
        render_algo_advisor(df_active)

    with eda_tab:
        render_eda_section(df_active)

    with run_tab:
        numeric_cols=df_active.select_dtypes(include=[np.number]).columns.tolist()
        all_cols=df_active.columns.tolist()

        for group_id,gmeta in {k:v for k,v in GROUP_META.items() if k!="balancing"}.items():
            methods_in=[(n,m) for n,m in METHODS.items() if m["group"]==group_id]
            st.markdown(f'<div style="font-size:.79rem;font-weight:700;color:{gmeta["color"]};margin:11px 0 6px;letter-spacing:.02em">{gmeta["icon"]} {gmeta["label"]}</div>',unsafe_allow_html=True)
            cols_m=st.columns(min(4,len(methods_in)))
            for i,(mname,meta) in enumerate(methods_in):
                in_sel=mname in st.session_state["selected_methods"]
                with cols_m[i%len(cols_m)]:
                    sel_cls=" sel" if in_sel else ""
                    sel_badge = '<span style="float:right;background:#10b981;color:white;font-size:.62rem;font-weight:700;padding:2px 7px;border-radius:10px">✅ Selected</span>' if in_sel else ""
                    name_color = "#34d399" if in_sel else "#f1f5f9"
                    st.markdown(
                        f'<div class="method-card{sel_cls}">{sel_badge}<span class="mbadge {meta["badge"]}">{group_id[:4].upper()}</span>'
                        f'<div class="method-name" style="color:{name_color}">{mname}</div><div class="method-vn">{meta["vn"]}</div>'
                        f'<div class="method-desc">{meta["desc"]}</div></div>',unsafe_allow_html=True)
                    if in_sel:
                        if st.button("✓ Remove",key=f"rm_{mname}"): st.session_state["selected_methods"].remove(mname); st.rerun()
                    else:
                        if st.button("＋ Select",key=f"add_{mname}"): st.session_state["selected_methods"].append(mname); st.rerun()

        sel=st.session_state["selected_methods"]
        if sel:
            sel_html=" ".join(f'<span style="background:rgba(59,130,246,.15);color:#60a5fa;padding:2px 9px;border-radius:4px;font-size:.74rem">{m}</span>' for m in sel)
            st.markdown(f'<div style="margin:7px 0">Selected: {sel_html}</div>',unsafe_allow_html=True)
            if st.button("🗑️ Clear all"): st.session_state["selected_methods"]=[]; st.rerun()

        st.markdown("---")
        has_clf=any(METHODS[m]["group"]=="classification" for m in sel)
        has_pred=any(METHODS[m]["group"]=="prediction" for m in sel)
        has_clus=any(METHODS[m]["group"]=="association" and "Cluster" in m for m in sel)
        has_assoc="Association Rules (Apriori)" in sel
        target_col=None; feature_cols=[]; test_size=.2; balance_opt="None"
        n_clusters=3; clus_features=[]; min_sup=.05; min_conf=.3; min_lift=1.0

        if has_clf or has_pred:
            sec_hdr("⚙️","Configure Supervised Learning","blue")
            ca2,cb2=st.columns(2)
            with ca2:
                auto_t=suggest_target(df_active)
                if auto_t: st.markdown(f'<div style="font-size:.74rem;color:#34d399;margin-bottom:3px">✅ Auto-suggested: <b>{auto_t}</b></div>',unsafe_allow_html=True)
                def_idx=all_cols.index(auto_t) if auto_t and auto_t in all_cols else 0
                target_col=st.selectbox("🎯 Target column",all_cols,index=def_idx)
            with cb2:
                feature_cols=st.multiselect("📐 Feature columns",[c for c in all_cols if c!=target_col],
                    default=[c for c in numeric_cols if c!=target_col][:10])
            _ts_c1, _ts_c2 = st.columns([4, 1])
            with _ts_c1:
                test_size_pct = st.slider(
                    "Validation split %", 5, 50,
                    st.session_state.get("_test_split_pct", 20),
                    step=1,
                    help="Kéo để chọn % dữ liệu dùng để validation. VD: 20% = train 80%, validation 20%.",
                    key="_ts_slider"
                )
            st.session_state["_test_split_pct"] = test_size_pct
            with _ts_c2:
                st.markdown(
                    f'<div style="background:rgba(59,130,246,.15);border:1px solid rgba(59,130,246,.3);'
                    f'border-radius:8px;padding:8px 12px;text-align:center;margin-top:4px">'
                    f'<div style="font-size:22px;font-weight:800;color:#60a5fa">{test_size_pct}%</div>'
                    f'<div style="font-size:10px;color:#6b7280">validation</div>'
                    f'<div style="font-size:10px;color:#34d399">{100-test_size_pct}% train</div>'
                    f'</div>', unsafe_allow_html=True
                )
            test_size = test_size_pct / 100
            if has_clf:
                balance_opt=st.selectbox("Class balancing",["None","Random Oversampling","SMOTE"],
                    help="None: không cân bằng | Oversampling: nhân bản dữ liệu thiểu số | SMOTE: tổng hợp điểm dữ liệu mới")
                if balance_opt != "None":
                    _y_check=df_active[target_col].dropna() if target_col else pd.Series()
                    if len(_y_check)>0:
                        _vc=_y_check.value_counts()
                        _ratio=_vc.min()/_vc.max() if _vc.max()>0 else 1
                        if _ratio > 0.8:
                            st.info(f"ℹ️ Data khá cân bằng (ratio={_ratio:.0%}) — balancing có thể không cần thiết.")
                        elif _ratio < 0.2:
                            st.warning(f"⚠️ Data rất mất cân bằng (ratio={_ratio:.0%}). {balance_opt} sẽ tăng minority class.")
                            if balance_opt == "SMOTE":
                                st.warning("🔴 **Cảnh báo SMOTE**: Dữ liệu rất mất cân bằng + SMOTE có nguy cơ **overfitting** — SMOTE tạo điểm tổng hợp dựa trên minority samples gốc, nếu tập quá nhỏ sẽ tạo ra overlap. Kiểm tra AUC trên validation set thực.")

        if has_clus:
            sec_hdr("⚙️","Configure Clustering","green")
            cc1,cc2=st.columns(2)
            with cc1: clus_features=st.multiselect("Clustering features",numeric_cols,default=numeric_cols[:8])
            with cc2: n_clusters=st.slider("Number of clusters (K)",2,10,3)

        if has_assoc:
            sec_hdr("⚙️","Configure Association Rules","amber")
            ra1,ra2,ra3=st.columns(3)
            with ra1: min_sup=st.slider("Min Support",0.01,.5,.05,.01)
            with ra2: min_conf=st.slider("Min Confidence",.1,1.,.3,.05)
            with ra3: min_lift=st.slider("Min Lift",1.,10.,1.,.1)

        if not sel: st.info("👆 Select at least one method above to run analysis.")
        elif st.button("🚀 Run Selected Methods",type="primary"):
            st.session_state["run_results"]=[]; st.session_state["run_exports"]={}
            st.session_state["_run_figures"]={}
            st.session_state["_fi_tables"]={}
            st.session_state["_trained_models"]={}
            st.session_state["_split_data"]={}
            for method in sel:
                group=METHODS[method]["group"]
                sec_hdr("▶","Running: "+method,"purple")
                try:
                    if group=="classification":
                        if not feature_cols: st.error(f"{method}: select feature columns."); continue
                        metrics,mdl,X_te,y_te,y_pred=run_classification(method,df_active,target_col,feature_cols,test_size,balance_opt)
                        st.session_state["run_results"].append(metrics)
                        pred_df=df_active.iloc[:len(y_te)][feature_cols].copy()
                        pred_df["Actual"]=y_te; pred_df["Predicted"]=y_pred
                        st.session_state["run_exports"][f"{method}_predictions"]=pred_df
                    elif group=="prediction":
                        if not feature_cols: st.error(f"{method}: select feature columns."); continue
                        metrics=run_regression(method,df_active,target_col,feature_cols,test_size)
                        st.session_state["run_results"].append(metrics)
                    elif group=="association":
                        if "Cluster" in method:
                            fc=clus_features if clus_features else numeric_cols[:6]
                            metrics,df_out=run_clustering(method,df_active,fc,n_clusters)
                            st.session_state["run_results"].append(metrics)
                            st.session_state["run_exports"][f"{method}_clusters"]=df_out
                        else:
                            metrics,rules_df=run_association(df_active,min_sup,min_conf,min_lift)
                            if metrics:
                                st.session_state["run_results"].append(metrics)
                                if rules_df is not None: st.session_state["run_exports"]["Association_Rules"]=rules_df
                    elif group=="balancing":
                        if not feature_cols: st.error(f"{method}: select feature columns."); continue
                        df_enc2=encode_df(df_active[feature_cols+[target_col]].dropna())
                        X_=df_enc2[feature_cols].values; y_=df_enc2[target_col].values
                        orig=pd.Series(y_).value_counts()
                        if method=="Random Oversampling": Xr,yr=RandomOverSampler(random_state=42).fit_resample(X_,y_)
                        else: Xr,yr=SMOTE(random_state=42).fit_resample(X_,y_)
                        bc1,bc2=st.columns(2)
                        with bc1: st.subheader("Before"); st.bar_chart(orig)
                        with bc2: st.subheader("After"); st.bar_chart(pd.Series(yr).value_counts())
                        st.success(f"{method}: {len(y_):,} → {len(yr):,} samples.")
                        st.session_state["run_results"].append({"Method":method,"Before rows":len(y_),"After rows":len(yr)})
                except Exception as e: st.error(f"Error running {method}: {e}")
            st.session_state["wizard_step"]=4; st.rerun()


    if ws==3:
        st.markdown("<br>",unsafe_allow_html=True)
        c1,_=st.columns([1,4])
        with c1:
            if st.button("Skip to Advise Mode →"): st.session_state["wizard_step"]=4; st.rerun()
    st.divider()

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 — ADVISE & PREDICT
# ═══════════════════════════════════════════════════════════════════════════════
if ws>=4:
    stack4=st.session_state["prep_transforms"].get(active_name,[])
    df_active=_json_to_df(stack4[-1][1]) if stack4 else raw_df.copy()
    sec_hdr("🧠","Step 4 — Advise & Predict","purple",subtitle="AI Blueprint · Results · Export")

    adv_tab1,adv_tab2,adv_tab3=st.tabs(["🤖 AI Blueprint","📊 Results & Analysis","⬇️ Export"])
    adv_tab4=None  # removed — dùng tab AI Chat ở top navigation

    with adv_tab1:
        sec_hdr("🎯","Describe Your Goal","blue",subtitle="AI will advise the best approach")
        user_goal=st.text_area("What do you want to predict or discover?",
            value=st.session_state.get("user_goal",""),
            placeholder="e.g. 'Predict which customers will churn' · 'Find customer segments' · 'Discover which products are bought together'…",
            height=80,key="ugi")
        st.session_state["user_goal"]=user_goal
        g_key=st.session_state.get("gemini_key","").strip(); o_key=st.session_state.get("openrouter_key","").strip()
        if not g_key and not o_key: st.warning("⚠️ Add an AI key in the sidebar. ML runs without one.")
        col_b1,col_b2,_=st.columns([1,1,3])
        with col_b1: analyse_clicked=st.button("🔎 Generate Blueprint",type="primary")
        with col_b2:
            if st.session_state["ai_blueprint"]:
                if st.button("🔄 Re-run"): st.session_state["ai_blueprint"]=""; st.rerun()

        if analyse_clicked and user_goal.strip():

            all_l4=st.session_state.get("sheets",{})
            summary_parts=[]
            for sn,sdf in all_l4.items():
                ps=st.session_state["prep_transforms"].get(sn,[])
                udf=_json_to_df(ps[-1][1]) if ps else sdf
                summary_parts.append(f"{'='*50}\nFILE: {sn}{'[CLEANED]' if ps else ''}\n{'='*50}\n"+df_summary(udf,sn))
            col_freq=Counter(c for d in all_l4.values() for c in d.columns)
            shared=[c for c,n in col_freq.items() if n>1]
            rel_note=f"\nSHARED COLUMNS (join keys): {', '.join(shared)}\n" if shared else "\nNo shared columns.\n"
            prompt=(f"You are an expert data scientist. {len(all_l4)} file(s) uploaded.\n\n"
                    +"\n\n".join(summary_parts)+rel_note+
                    f"\nUSER GOAL: {user_goal}\n\n"
                    "Provide a CONCISE blueprint:\n\n"
                    "1. DATASET UNDERSTANDING — file contents, column roles, data types.\n"
                    "2. DATA QUALITY ISSUES — per-file: missing %, outliers, text columns, duplicates.\n"
                    "3. RECOMMENDED TRANSFORMATIONS — fix order, merge strategy, feature engineering.\n"
                    "4. BEST TECHNIQUES — recommend 2-3 from: Logistic Regression, Random Forest, Linear Regression, Neural Networks (MLP), K-Means Clustering, Hierarchical Clustering, Association Rules (Apriori), LDA, KNN, Naive Bayes, SVM. Explain WHY. State top 2 to compare.\n"
                    "5. TARGET & FEATURE COLUMNS — exact column names, why this target.\n"
                    "6. EXPECTED CHALLENGES — class imbalance, high cardinality, collinearity.\n\n"
                    "FORMATTING: No **, no #, no dashes. Plain numbered paragraphs. Reference actual column names. Max 550 words.")
            with st.spinner(f"AI analysing {len(all_l4)} file(s)…"):
                result=ask_ai(prompt)
            st.session_state["ai_blueprint"]=result

        if st.session_state["ai_blueprint"]:
            render_ai_box(st.session_state["ai_blueprint"],"🤖 AI Blueprint")
            c_vn,_=st.columns([1,3])
            with c_vn:
                if st.button("🇻🇳 Dịch sang Tiếng Việt",key="tbp"):
                    with st.spinner("Đang dịch…"):
                        vn=ask_ai(f"Translate to natural Vietnamese. No markdown. Plain paragraphs.\n\n{st.session_state['ai_blueprint'][:3000]}")
                    st.session_state["ai_vn"]=vn
            if st.session_state.get("ai_vn"): render_ai_box(st.session_state["ai_vn"],"🇻🇳 Phân tích Tiếng Việt")
        if st.session_state["ai_blueprint"] and not st.session_state["run_results"]:
            st.markdown('<div style="background:rgba(16,185,129,.07);border:1px solid rgba(16,185,129,.25);border-radius:10px;padding:12px 16px;font-size:.82rem;color:#6ee7b7;margin-top:12px">💡 <b>Next:</b> Go to <b>Step 3 → Run ML Models</b> to execute recommended methods, then return here for full analysis.</div>',unsafe_allow_html=True)

    with adv_tab2:
        results=st.session_state.get("run_results",[])
        if not results: st.info("No results yet — run models in Step 3 first.")
        else:
            sec_hdr("📋","Methods Comparison","blue")
            cmp_df=pd.DataFrame(results)
            col_order=["Sheet","Method","Accuracy","Precision","Recall","F1-Score","AUC","R²","RMSE","K","Silhouette","Rules found","Train rows","Validation rows","Rows"]
            col_order=[c for c in col_order if c in cmp_df.columns]
            cmp_df=cmp_df[col_order]; st.dataframe(_sanitize_df(cmp_df),width='stretch')
            if "F1-Score" in cmp_df.columns:
                try:
                    best=cmp_df.loc[cmp_df["F1-Score"].astype(float).idxmax()]
                    st.success(f"🏆 Best: **{best.get('Method','')}** — F1 = {best['F1-Score']}")
                except Exception: pass
            elif "R²" in cmp_df.columns:
                try:
                    best=cmp_df.loc[cmp_df["R²"].astype(float).idxmax()]
                    st.success(f"🏆 Best regression: **{best.get('Method','')}** — R² = {best['R²']}")
                except Exception: pass

            render_fitness_report(results, df_active, st.session_state.get("user_goal",""))

            # ── Hiển thị lại các biểu đồ đã lưu khi chạy model ─────────────
            saved_figs = st.session_state.get("_run_figures", {})
            if saved_figs:
                st.markdown("---")
                sec_hdr("📊","Biểu đồ kết quả model","blue")
                for method_name, fig_bytes_list in saved_figs.items():
                    st.markdown(
                        f'<div style="font-size:.85rem;font-weight:700;color:#60a5fa;'
                        f'margin:14px 0 8px;border-left:3px solid #3b82f6;padding-left:10px">'
                        f'▶ {method_name}</div>', unsafe_allow_html=True)
                    # Chia đều các figure vào cột (tối đa 3/hàng)
                    n = len(fig_bytes_list)
                    cols = st.columns(min(n, 3))
                    for i, img_bytes in enumerate(fig_bytes_list):
                        with cols[i % min(n, 3)]:
                            st.image(img_bytes, use_container_width=True)

            # ── Export PNG biểu đồ ─────────────────────────────────────────
            saved_figs = st.session_state.get("_run_figures", {})
            if saved_figs:
                sec_hdr("📥","Tải xuống biểu đồ","green")
                for mname, fig_list in saved_figs.items():
                    for i, img_bytes in enumerate(fig_list):
                        chart_labels = ["Confusion_Matrix","ROC_Curve","Feature_Importance","Actual_vs_Predicted","Residuals","Elbow_Chart","Cluster_Plot"]
                        label = chart_labels[i] if i < len(chart_labels) else f"Chart_{i+1}"
                        st.download_button(f"⬇️ {mname} — {label}.png",
                            data=img_bytes, file_name=f"{mname}_{label}.png",
                            mime="image/png", key=f"dl_fig_{mname}_{i}")

            # ── Feature Importance Table ───────────────────────────────────
            fi_tables = st.session_state.get("_fi_tables", {})
            if fi_tables:
                sec_hdr("📊","Feature Importance","blue")
                for mname, fi_df in fi_tables.items():
                    with st.expander(f"📋 {mname} — Top {len(fi_df)} Features", expanded=False):
                        st.dataframe(_sanitize_df(fi_df), width='stretch')
                        fi_csv = fi_df.to_csv(index=False).encode()
                        st.download_button(f"⬇️ Download {mname}_features.csv",
                            data=fi_csv, file_name=f"{mname}_feature_importance.csv",
                            mime="text/csv", key=f"dl_fi_{mname}")

            st.markdown("---")
            sec_hdr("🔮","Predict New Data","purple",subtitle="Download template, nhap data moi, upload de du bao")
            trained = st.session_state.get("_trained_models",{})
            if not trained:
                st.info("Chua co model nao duoc train. Chay model o Step 3 truoc.")
            else:
                _pm_col, _pu_col = st.columns([2,3])
                with _pm_col:
                    pred_model_name = st.selectbox("Chon model de du bao",list(trained.keys()),key="pred_model_sel")
                with _pu_col:
                    # Download template
                    _mdl_info = trained[pred_model_name]
                    _tpl_df = pd.DataFrame(columns=_mdl_info["features"])
                    _tpl_df.loc[0] = [""] * len(_mdl_info["features"])
                    _tpl_buf = io.BytesIO()
                    with pd.ExcelWriter(_tpl_buf, engine="openpyxl") as _w:
                        _tpl_df.to_excel(_w, sheet_name="Input", index=False)
                    st.download_button(
                        "Download Template (Excel)",
                        data=_tpl_buf.getvalue(),
                        file_name=f"template_{pred_model_name[:20].replace(' ','_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_template"
                    )

                pred_file = st.file_uploader("Upload data moi de du bao (.xlsx / .csv)",
                    type=["xlsx","csv"], key="pred_upload")

                if pred_file:
                    try:
                        if pred_file.name.endswith(".csv"):
                            _new_df = pd.read_csv(pred_file)
                        else:
                            _new_df = pd.read_excel(pred_file)

                        _feats = _mdl_info["features"]
                        _missing_cols = [c for c in _feats if c not in _new_df.columns]
                        if _missing_cols:
                            st.error(f"Thieu cot: {_missing_cols}. Vui long dung dung template.")
                        else:
                            _X_new = _mdl_info["scaler"].transform(_new_df[_feats].fillna(0).values)
                            _preds = _mdl_info["model"].predict(_X_new)
                            _result_df = _new_df[_feats].copy()
                            _result_df["Prediction"] = _preds
                            if _mdl_info["type"] == "classification" and hasattr(_mdl_info["model"],"predict_proba"):
                                _proba = _mdl_info["model"].predict_proba(_X_new)
                                _result_df["Confidence"] = _proba.max(axis=1).round(4)
                            st.success(f"Du bao {len(_result_df)} dong thanh cong!")
                            st.dataframe(_sanitize_df(_result_df), width='stretch')
                            # Download result
                            _res_buf = io.BytesIO()
                            with pd.ExcelWriter(_res_buf, engine="openpyxl") as _rw:
                                _sanitize_df(_result_df).to_excel(_rw, sheet_name="Predictions", index=False)
                            st.download_button(
                                "Download Ket qua Du bao (Excel)",
                                data=_res_buf.getvalue(),
                                file_name=f"predictions_{pred_model_name[:20].replace(' ','_')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_predictions"
                            )
                    except Exception as _pe:
                        st.error(f"Loi du bao: {_pe}")


            st.markdown("---"); sec_hdr("🤖","AI Analysis of Results","purple")
            if st.button("🔎 Generate AI Explanation",type="primary",key="gaie"):
                results_text=cmp_df.to_string(index=False)
                prompt2=(f"You are a data mining expert.\n\nDATASET: {active_name} ({df_active.shape[0]} rows x {df_active.shape[1]} cols)\n"
                         f"USER GOAL: {st.session_state.get('user_goal','(not specified)')}\n\nRESULTS:\n{results_text}\n\n"
                         "Provide:\n1. OVERALL PERFORMANCE SUMMARY — best/worst and why.\n"
                         "2. METHOD-BY-METHOD ANALYSIS — what each metric means.\n"
                         "3. COMPARISON INSIGHTS — winner and what differences reveal.\n"
                         "4. PRACTICAL RECOMMENDATIONS — what to deploy, what to improve.\n"
                         "5. RED FLAGS — overfitting, class imbalance, low AUC, weak silhouette.\n\n"
                         "FORMATTING: No **, no #, no dashes. Plain numbered paragraphs. Max 500 words.")
                with st.spinner("AI analysing results…"):
                    ai_exp=ask_ai(prompt2)
                st.session_state["comparison_ai"]=ai_exp
            if st.session_state.get("comparison_ai"):
                render_ai_box(st.session_state["comparison_ai"],"🤖 AI Analysis")
                c_vn2,_=st.columns([1,3])
                with c_vn2:
                    if st.button("🇻🇳 Dịch kết quả",key="trr"):
                        with st.spinner("Đang dịch…"):
                            txt2=st.session_state["comparison_ai"][:3000]
                            vn2=ask_ai("Translate to natural Vietnamese. No markdown. "+txt2)
                        st.session_state["ai_vn"]=vn2
                if st.session_state.get("ai_vn"): render_ai_box(st.session_state["ai_vn"],"🇻🇳 Giải thích Tiếng Việt")

    with adv_tab3:
        sec_hdr("⬇️","Export All Results","green")
        results_exp=st.session_state.get("run_results",[])
        if not results_exp: st.info("Run models in Step 3 first.")
        else:
            cmp_df2=pd.DataFrame(results_exp)
            exp_sheets={}

            # \u2500\u2500 Sheet 1: Raw Data (Before) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
            raw_sheet = st.session_state["sheets"].get(active_name, raw_df)
            exp_sheets["Raw Data (Before)"] = _sanitize_df(raw_sheet.copy())

            # \u2500\u2500 Sheet 2: Cleaned Data (After) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
            _prep_stack4 = st.session_state["prep_transforms"].get(active_name, [])
            if _prep_stack4:
                cleaned_df = _json_to_df(_prep_stack4[-1][1])
                exp_sheets["Cleaned Data (After)"] = _sanitize_df(cleaned_df)
            else:
                exp_sheets["Cleaned Data (same as raw)"] = _sanitize_df(raw_sheet.copy())

            # \u2500\u2500 Sheet 3+: K\u1ebft qu\u1ea3 model \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
            # -- Train / Validation data per model
            for mname, split in st.session_state.get("_split_data",{}).items():
                short = mname[:20].replace(" ","_")
                exp_sheets[f"Train_{short}"] = _sanitize_df(split["train"].copy())
                exp_sheets[f"Val_{short}"] = _sanitize_df(split["validation"].copy())

            exp_sheets["Comparison Table"] = cmp_df2
            for label,df_exp in st.session_state.get("run_exports",{}).items():
                exp_sheets[label[:31]]=df_exp
            if st.session_state.get("comparison_ai"):
                exp_sheets["AI Explanation"]=pd.DataFrame({"AI Explanation":st.session_state["comparison_ai"].split("\n")})
            if st.session_state.get("ai_blueprint"):
                exp_sheets["AI Blueprint"]=pd.DataFrame({"AI Blueprint":st.session_state["ai_blueprint"].split("\n")})
            for mname, fi_df in st.session_state.get("_fi_tables",{}).items():
                exp_sheets[f"FI_{mname[:22]}"]=fi_df

            # Preview c\u00e1c sheet s\u1ebd xu\u1ea5t
            for sname, df_p in exp_sheets.items():
                _clr = "#10b981" if ("Cleaned" in sname or "After" in sname) else ("#ef4444" if ("Raw" in sname or "Before" in sname) else "#60a5fa")
                _tag = "CLEAN" if ("Cleaned" in sname or "After" in sname) else ("RAW" if ("Raw" in sname or "Before" in sname) else "DATA")
                _info = f"{df_p.shape[0]:,} rows x {df_p.shape[1]} cols"
                st.markdown(
                    f'<div style="background:#0f1629;border:1px solid #1f2937;border-radius:7px;padding:6px 12px;margin-bottom:4px;font-size:.79rem;color:#9ca3af">'
                    f'<span style="color:{_clr};font-weight:700;margin-right:6px">[{_tag}]</span>'
                    f'<b style="color:#60a5fa">{sname}</b> - {_info}</div>',
                    unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)
            figures_dict = st.session_state.get("_run_figures", {})
            excel_bytes=export_to_excel(exp_sheets, figures_dict=figures_dict if figures_dict else None)
            st.download_button("Download Full Results + Charts (Excel)",data=excel_bytes,
                file_name=f"DataMineAI_{active_name[:20].replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("---")
    if st.button("Start New Analysis (reset all)"):
        for key2 in ["wizard_step","ai_blueprint","prep_transforms","prep_log","enc_mapping",
                     "user_goal","selected_methods","run_results","run_exports","comparison_ai","ai_vn","chat_messages","_run_figures","_fi_tables","_trained_models","_split_data"]:
            if isinstance(DEFAULTS.get(key2),dict): st.session_state[key2]={}
            elif isinstance(DEFAULTS.get(key2),list): st.session_state[key2]=[]
            else: st.session_state[key2]=DEFAULTS.get(key2,"")
        st.session_state["wizard_step"]=1; st.rerun()
