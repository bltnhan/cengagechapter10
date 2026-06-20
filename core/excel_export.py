"""
core.excel_export — Xuất kết quả ra Excel (openpyxl), thuần Python.

Tách từ Streamlit.py: _fmt_score_sheet / export_to_excel
Logic GIỮ NGUYÊN — các hàm này vốn đã không phụ thuộc Streamlit.
"""
from __future__ import annotations

import io

import pandas as pd


def _fmt_score_sheet(ws, sheet_name: str = ""):
    """Apply Solver Analytics-style formatting to TrainScore / ValScore sheets."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Color palette
    BLUE_HDR = PatternFill("solid", fgColor="1E3A5F")   # section header
    BLUE_COL = PatternFill("solid", fgColor="1D4ED8")   # column header
    GREEN_H = PatternFill("solid", fgColor="064E3B")    # green highlight row
    AMBER_H = PatternFill("solid", fgColor="451A03")    # amber section
    LIGHT_ROW = PatternFill("solid", fgColor="0F172A")  # data row background
    ALT_ROW = PatternFill("solid", fgColor="111827")    # alternating row
    WHITE = Font(color="F1F5F9", bold=True, name="Calibri", size=10)
    BOLD_W = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    DATA_F = Font(color="E2E8F0", name="Calibri", size=10)
    GREEN_F = Font(color="34D399", bold=True, name="Calibri", size=10)
    BLUE_F = Font(color="60A5FA", bold=True, name="Calibri", size=11)
    thin = Side(style="thin", color="1E293B")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rgt = Alignment(horizontal="right", vertical="center")

    SEC_KEYS = {"Confusion Matrix", "Error Report", "Metrics", "Validation: Classification Details"}
    COL_KEYS = {"Actual\\Predicted", "Actual/Predicted", "Class", "Metric", "Record ID"}
    METRIC_KEYS = {"Accuracy (#correct)", "Accuracy (%correct)", "Specificity",
                   "Sensitivity (Recall)", "Precision", "F1 score", "AUC (ROC)"}

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if cell.value is None:
                continue
            v = str(cell.value).strip()
            col = cell.column

            if v in SEC_KEYS:
                cell.fill = BLUE_HDR; cell.font = BOLD_W; cell.alignment = lft
            elif v in COL_KEYS or (col == 1 and v in ("0", "1", "Overall")):
                cell.fill = BLUE_COL; cell.font = WHITE; cell.alignment = ctr
            elif col == 1 and v in METRIC_KEYS:
                cell.fill = ALT_ROW; cell.font = DATA_F; cell.alignment = lft
            elif col == 2 and row[0].value in METRIC_KEYS:
                cell.fill = ALT_ROW; cell.font = BLUE_F; cell.alignment = rgt
                try:
                    fv = float(v)
                    if 0 < fv <= 1:
                        cell.number_format = "0.0000"
                    elif fv > 1 and fv <= 100:
                        cell.number_format = "0.00"
                except Exception:
                    pass
            elif col == 3 and row[0].value in METRIC_KEYS:
                cell.fill = ALT_ROW; cell.font = Font(color="94A3B8", italic=True, size=9); cell.alignment = lft
            else:
                cell.fill = LIGHT_ROW if row_idx % 2 == 0 else ALT_ROW
                cell.font = DATA_F; cell.alignment = lft
            cell.border = bdr

    col_widths = {1: 28, 2: 16, 3: 40}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 20

    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Data Science: Classification Report — {sheet_name}"
    title_cell.fill = PatternFill("solid", fgColor="0F172A")
    title_cell.font = Font(color="60A5FA", bold=True, size=12, name="Calibri")
    title_cell.alignment = lft
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.row_dimensions[1].height = 24


def export_to_excel(results_dict: dict, figures_dict: dict | None = None) -> bytes:
    """
    Xuất Excel với formatting đẹp theo chuẩn Solver Analytics.

    Args:
        results_dict : {sheet_name: DataFrame}
        figures_dict : {method_name: [png_bytes, ...]} — nhúng vào sheet "Charts".

    Returns:
        bytes của file .xlsx.
    """
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in results_dict.items():
            df.to_excel(writer, sheet_name=str(name)[:31], index=False)
            try:
                _fmt_score_sheet(writer.sheets[str(name)[:31]], sheet_name=name)
            except Exception:
                pass

        if figures_dict:
            wb = writer.book
            if "Charts" not in wb.sheetnames:
                ws_charts = wb.create_sheet("Charts")
            else:
                ws_charts = wb["Charts"]

            row_offset = 1
            for method_name, fig_bytes_list in figures_dict.items():
                chart_labels = ["Confusion_Matrix", "ROC_Curve", "Feature_Importance",
                                "Actual_vs_Predicted", "Residuals", "Elbow_Chart", "Cluster_Plot"]
                ws_charts.cell(row=row_offset, column=1, value=f"▶ {method_name}")
                ws_charts.cell(row=row_offset, column=1).font = Font(bold=True, size=12, color="3B82F6")
                row_offset += 1

                col_offset = 1
                for i, img_bytes in enumerate(fig_bytes_list):
                    try:
                        img_io = io.BytesIO(img_bytes)
                        xl_img = XLImage(img_io)
                        xl_img.width = 380
                        xl_img.height = 260
                        col_letter = chr(ord("A") + (col_offset - 1) * 7)
                        cell_ref = f"{col_letter}{row_offset}"
                        ws_charts.add_image(xl_img, cell_ref)
                        label = chart_labels[i] if i < len(chart_labels) else f"Chart_{i + 1}"
                        ws_charts.cell(
                            row=row_offset + 15,
                            column=(col_offset - 1) * 7 + 1,
                            value=label.replace("_", " "),
                        )
                        col_offset += 1
                    except Exception:
                        pass
                row_offset += 19

    return buf.getvalue()
